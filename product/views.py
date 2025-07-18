from django.forms.models import BaseModelForm
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.utils import IntegrityError
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from user.permission import IsAdminMixin
from .models import ProductCategory
from .forms import ProductCategoryForm
from bill.utils import update_subledger_after_updating_product

class ProductCategoryMixin(IsAdminMixin):
    model = ProductCategory
    form_class = ProductCategoryForm
    paginate_by = 50
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_category_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class ProductCategoryList(ProductCategoryMixin, ListView):
    template_name = "productcategory/productcategory_list.html"
    queryset = ProductCategory.objects.filter(status=True, is_deleted=False)


class ProductCategoryDetail(ProductCategoryMixin, DetailView):
    template_name = "productcategory/productcategory_detail.html"


class ProductCategoryCreate(ProductCategoryMixin, CreateView):
    template_name = "create.html"


class ProductCategoryUpdate(ProductCategoryMixin, UpdateView):
    template_name = "update.html"


class ProductCategoryDelete(ProductCategoryMixin, DeleteMixin, View):
    pass


from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import Product
from .forms import ProductForm


class ProductMixin(IsAdminMixin):
    model = Product
    form_class = ProductForm
    paginate_by = 50
    queryset = Product.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("product_list")
    search_lookup_fields = [
        "title",
        "description",
    ]


class ProductList(ProductMixin, ListView):
    template_name = "product/product_list.html"
    queryset = Product.objects.filter(status=True, is_deleted=False)


class ProductDetail(ProductMixin, DetailView):
    template_name = "product/product_detail.html"


class ProductCreate(ProductMixin, CreateView):
    template_name = "create.html"


class ProductUpdate(ProductMixin, UpdateView):
    template_name = "update.html"

    def form_valid(self, form):
        updated_name = form.data.get('title')
        product_id = form.initial.get('id')
        initial_name = form.initial.get('title')
        update_subledger_after_updating_product(product_id=product_id, initial_name=initial_name, updated_name=updated_name)
        return super().form_valid(form)

import shortuuid
class ProductDelete(ProductMixin, View):
    # pass
    def remove_from_DB(self, request):
        try:
            object_id = request.GET.get("pk", None)
            product = self.model.objects.get(id=object_id)
            product.slug = f"deleted-{str(shortuuid.uuid()[:12])}"
            product.is_deleted = True
            product.status = False

            # Change slug to avoid duplicates on future inserts

            print(f"product slug changed deleted-{shortuuid.uuid()[:4]}")
            product.save()

            return True
        except Exception as e:
            print(e)
            return str(e)

    def get(self, request):
        status = self.remove_from_DB(request)
        return JsonResponse({"deleted": status})

from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import CustomerProduct
from .forms import CustomerProductForm


class CustomerProductMixin(IsAdminMixin):
    model = CustomerProduct
    form_class = CustomerProductForm
    paginate_by = 50
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("customerproduct_list")
    search_lookup_fields = ["product__title", "customer__name", "agent__full_name"]


class CustomerProductList(CustomerProductMixin, ListView):
    template_name = "customerproduct/customerproduct_list.html"
    queryset = CustomerProduct.objects.filter(status=True, is_deleted=False)


class CustomerProductDetail(CustomerProductMixin, DetailView):
    template_name = "customerproduct/customerproduct_detail.html"


class CustomerProductCreate(CustomerProductMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductUpdate(CustomerProductMixin, UpdateView):
    template_name = "update.html"

    def form_valid(self, form):
        form.instance.agent = self.request.user
        return super().form_valid(form)


class CustomerProductDelete(CustomerProductMixin, DeleteMixin, View):
    pass

'''  STock VIews '''

from .models import ProductStock
from .forms import ProductStockForm

class ProductStockMixin:
    model = ProductStock
    form_class = ProductStockForm
    paginate_by = 10
    queryset = ProductStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('productstock_list')


class ProductStockList(ProductStockMixin, ListView):
    template_name = "productstock/productstock_list.html"
    queryset = ProductStock.objects.filter(status=True,is_deleted=False)

class ProductStockDetail(ProductStockMixin, DetailView):
    template_name = "productstock/productstock_detail.html"

class ProductStockCreate(ProductStockMixin, CreateView):
    template_name = "create.html"

class ProductStockUpdate(ProductStockMixin, UpdateView):
    template_name = "update.html"

class ProductStockDelete(ProductStockMixin, DeleteMixin, View):
    pass


from openpyxl import load_workbook
from django.conf import settings


class ProductUpload(View):

    def post(self, request):
        file = request.FILES['file']
        wb = load_workbook(file)

        excel_data = list()
        for sheet in wb.worksheets:
            # Define the starting row where your data begins
            start_row = 2  # Assuming data starts from row 2 (1-indexed)

            for row in sheet.iter_rows(min_row=start_row):
                row_data = [cell.value for cell in row]
                
                # Check if the row contains valid data
                if row_data[0] and row_data[1] and row_data[2] and row_data[3] and row_data[4]:
                    excel_data.append(row_data)
     

        product_create_error = []
        for data in excel_data:
            # if not all(data):
            #     continue
            if data[0].strip().lower().startswith('category'):
                continue
            category_name = data[0].strip().lower().title()

            if ProductCategory.objects.filter(title__iexact=category_name).exists():
                category = ProductCategory.objects.get(title__iexact=category_name)
            else:
                try:
                    category = ProductCategory.objects.create(title=category_name)
                except IntegrityError:
                    category = ProductCategory.objects.get(title__iexact=category_name)
            
            product = Product(category=category, title=data[1].strip(), price=float(data[2]), unit=data[3])
            product.is_taxable = True if data[4].strip().lower() == "yes" else False

            try:
                product.save()
            except Exception as e:
                print(e)
                product_create_error.append(product.title)

        if product_create_error:
            messages.error(request, f"Error creating products \n {product_create_error}", extra_tags='danger')
            return redirect(reverse_lazy('product_list'))

        messages.success(request, "Products uploaded successfully", extra_tags='success')
        return redirect(reverse_lazy('product_list'))
    

from .models import BranchStock, BranchStockTracking, ItemReconcilationApiItem
from organization.models import Branch, EndDayRecord
from .forms import BranchStockForm
class BranchStockMixin(IsAdminMixin):
    model = BranchStock
    form_class = BranchStockForm
    paginate_by = 10
    queryset = BranchStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('branchstock_list')

    search_lookup_fields = [
        "product__title",
        "product__description",
    ]

class BranchStockList(BranchStockMixin, ListView):
    template_name = "branchstock/branchstock_list.html"

class BranchStockDetail(BranchStockMixin, DetailView):
    template_name = "branchstock/branchstock_detail.html"

class BranchStockCreate(BranchStockMixin, CreateView):
    template_name = "branchstock/branchstock_create.html"

class BranchStockUpdate(BranchStockMixin, UpdateView):
    template_name = "update.html"

class BranchStockDelete(BranchStockMixin, DeleteMixin, View):
    pass

from django.db.models import Sum
from bill.models import Bill
from datetime import datetime, date
from .utils import check_opening_for_branch

class ReconcileView(View): 

    def get(self, request):
        branch_opening_status_list = check_opening_for_branch()

        opening_exists = BranchStockTracking.objects.count()
        if opening_exists <= 0:
            return render(request, 'item_reconcilation/reconcilation.html',{'show_opening':True})

        branch = request.GET.get('branch', None)
        filter_date = request.GET.get('date')
        branches = Branch.objects.filter(is_deleted=False)
        if not branch or not filter_date:
            return render(request, 'item_reconcilation/reconcilation.html',{'message':'Please Select a Branch and Date', 'branches':branches, 'branch_status_list':branch_opening_status_list})
        try:
            filter_date = datetime.strptime(filter_date, '%Y-%m-%d').date().strftime('%Y-%m-%d')
        except Exception:
            return render(request, 'item_reconcilation/reconcilation.html',{'message':'Date format must me YYYY-mm-dd', 'branches':branches, 'branch_status_list':branch_opening_status_list})

        filter_branch = get_object_or_404(Branch, branch_code__iexact=branch)

        exists_in_bst = BranchStockTracking.objects.filter(date=filter_date, branch=filter_branch).exists()
        if not exists_in_bst:
            products = Product.objects.filter(reconcile=True).order_by('title').values()
            api_items = ItemReconcilationApiItem.objects.filter(date=filter_date, branch=filter_branch, product__reconcile=True).values()
            received = RequisitionBranchStock.objects.filter(created_at__contains=filter_date, branch=filter_branch, product__reconcile=True).values('product').annotate(quantity=Sum('quantity'))
            bills = Bill.objects.filter(transaction_date=filter_date, branch=filter_branch, status=True)

            new_products = {}
            for product in products:
                for k, v in product.items():
                    if k =='id':
                        physical_count = 0
                        if opening_exists > 0:
                            if BranchStockTracking.objects.filter(branch=filter_branch, product_id=v).exists():
                                physical_count = BranchStockTracking.objects.filter(branch=filter_branch, product_id=v).last().physical
                        new_products[str(v)] = {'title':product.get('title'), 'opening': physical_count}
                        break
        
            for item in api_items:
                product_id = str(item.get('product_id'))
                new_products[product_id]['wastage'] = item.get('wastage', 0)
                new_products[product_id]['returned'] = item.get('returned', 0)
                new_products[product_id]['physical'] = item.get('physical', 0)

            for rec in received:
                product_id = str(rec.get('product'))
                new_products[product_id]['received'] = rec.get('quantity')
            
            for bill in bills:
                for item in bill.bill_items.all():
                    product_id = str(item.product.id)
                    if item.product.reconcile:
                        has_sold = new_products[product_id].get('sold', None)
                        if has_sold:
                            new_products[product_id]['sold'] += item.product_quantity
                        else:
                            new_products[product_id]['sold'] = item.product_quantity

            product_to_view = []
            for k,v in new_products.items():
                new_dict = {'id': k, **v}
                if not 'opening' in new_dict:
                    new_dict['opening'] = 0
                if not 'received' in new_dict:
                    new_dict['received'] = 0
                if not 'wastage' in new_dict:
                    new_dict['wastage'] = 0
                if not 'returned' in new_dict:
                    new_dict['returned'] = 0
                if not 'sold' in new_dict:
                    new_dict['sold'] = 0
                if not 'closing' in new_dict:
                    new_dict['closing'] = 0
                if not 'physical' in new_dict:
                    new_dict['physical'] = 0
                if not 'discrepancy' in new_dict:
                    new_dict['discrepancy'] = 0

                product_to_view.append(new_dict)
            
            for prd in product_to_view:
                opening_received = prd.get('opening') + prd.get('received')
                wastage_returned_sold = prd.get('wastage') + prd.get('returned') + prd.get('sold')
                closing_value = opening_received - wastage_returned_sold
                prd['closing'] = closing_value
                prd['discrepancy'] = prd.get('physical') - closing_value

            context = {
                'products':product_to_view,
                'branches':branches,
                'should_save':True,
                'opening_exists': opening_exists,
                'branch_status_list':branch_opening_status_list
            }
            return render(request, 'item_reconcilation/reconcilation.html',context)
        
        # --------------------------

        products = BranchStockTracking.objects.filter(date=filter_date, branch=filter_branch).order_by('product__title')
        context = {
            'products':products,
            'branches':branches,
            'should_save':False,
            'opening_exists': opening_exists,
            'branch_status_list':branch_opening_status_list
        }
        return render(request, 'item_reconcilation/reconcilation.html', context)
    
    def post(self, request):
        branches = Branch.objects.all()
        branch_code = request.POST.get('branch').lower()
        reconcile_date = request.POST.get('filter_date')
        branch = get_object_or_404(Branch, branch_code__iexact=branch_code)
        today_date = date.today()
        if datetime.strptime(reconcile_date, '%Y-%m-%d').date() > today_date:
            messages.error(request, f"Date must not be greater than {today_date}")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})
        
        if BranchStockTracking.objects.filter(date__gte=reconcile_date, branch=branch).exists():
            messages.error(request, f"Items from date greater than {reconcile_date} exists")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})

        last_bill_in_tracking_date = BranchStockTracking.objects.last().date
        bill_exists = Bill.objects.filter(transaction_date__gt=last_bill_in_tracking_date, transaction_date__lt= reconcile_date ,status=True).exists()
        api_items_exists = ItemReconcilationApiItem.objects.filter(date__gt=last_bill_in_tracking_date,date__lt=reconcile_date).exists()

        if bill_exists or api_items_exists:
            messages.error(request, f"Please reconcile items form previous date/s")
            return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})

        data = request.POST
        for k in data:
            try:
                product_id = int(k)
                details = data.getlist(k)
                BranchStockTracking.objects.create(
                    product_id=product_id,
                    branch=branch,
                    date=reconcile_date,
                    opening=details[0],
                    received=details[1],
                    wastage=details[2],
                    returned=details[3],
                    sold=details[4],
                    closing=details[5],
                    physical=details[6],
                    discrepancy=details[7],
                    )
            except ValueError:
                pass
            except IntegrityError:
                messages.error(request, "Items for Today's date already exists")
                return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})
        return render(request, 'item_reconcilation/reconcilation.html', {'branches':branches})


class BranchStockUploadView(View):
    
    def post(self, request):
        # if BranchStockTracking.objects.count() > 0:
        #     messages.error(request, "Opening data already exists!!")
        #     return redirect(reverse_lazy("reconcile"))
        file = request.FILES.get('file')
        branches = Branch.objects.all()
        branch_dict = {}
        for b in branches:
            branch_dict[b.branch_code.lower()] = b.pk

        wb = load_workbook(file)
        excel_data = list()
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                row_data = list()
                for cell in row:
                    if cell.value:
                        row_data.append(str(cell.value))
                row_data.append(sheet.title)
                excel_data.append(row_data)
       
        product_dict = {}
        for d in excel_data:
            if len(d) < 3:
                continue
            if d[0].lower().startswith('date'):
                continue
            try:
                product_title = d[1].lower().strip()
                product = product_dict.get(product_title, None)
                if not product:
                    product_dict[product_title] = Product.objects.get(title__iexact=product_title).pk
                product_id = product_dict.get(product_title)
                branch_id =  branch_dict.get(d[3].lower())
                quantity = int(d[2])
                opening_date = datetime.strptime(d[0][:10], '%Y-%m-%d')
                BranchStockTracking.objects.create(product_id=product_id, branch_id=branch_id, opening=quantity, physical=quantity, date=opening_date)
                BranchStock.objects.create(product_id=product_id,quantity=quantity, branch_id=branch_id)
            except Exception as e:
                print(e)

        return redirect(reverse_lazy("reconcile"))


class UpdateDateForReconcilationView(View):

    def post(self, request):
        from_date = request.POST.get('from_date', None)
        to_date = request.POST.get('to_date', None)

        if not from_date or not to_date:
            messages.error(request, 'Please Provide both "From date" and "To date"')
            return redirect('/reconcile')
        ItemReconcilationApiItem.objects.filter(date=from_date).update(date=to_date)
        EndDayRecord.objects.filter(date=from_date).update(date=to_date)
        messages.success(request, 'Date has been updated')
        return redirect('/reconcile')
        
from django.db.models import OuterRef, Subquery, Sum, F
from django.db.models.functions import Coalesce
from django.db.models import Q       
class BranchStockTotalList(View):
    def get(self, request):
        branch_id = request.GET.get('branch')  # Fetch branch_id from GET parameters
        branches = Branch.objects.all()

        branchstock_quantity_subquery_all = BranchStock.objects.filter(
            product=OuterRef('pk'),
            branch_id=branch_id,
            is_deleted=False

        ).values('product').annotate(total_quantity=Sum('quantity')).values('total_quantity')[:1]

        products = Product.objects.annotate(
            branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
        ).filter(
            is_deleted=False,
            branchstock_total_quantity__gt=0,
        )

        return render(request, 'branchstock/branchstock_total.html', {'branches': branches})

    def post(self, request):
        branch_id = request.POST.get('branch')
        branch = Branch.objects.get(pk=branch_id)
        search_query = request.POST.get('search')  # Fetch search query from POST parameters

        branchstock_quantity_subquery_all = BranchStock.objects.filter(
            product=OuterRef('pk'),
            branch=branch,
            is_deleted=False

        ).values('product').annotate(total_quantity=Sum('quantity')).values('total_quantity')[:1]

        if search_query:
            products = Product.objects.annotate(
                branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
            ).filter(
                title__icontains = search_query,
                is_deleted=False,
                branchstock_total_quantity__gt=0
            )
        else:
            products = Product.objects.annotate(
                branchstock_total_quantity=Subquery(branchstock_quantity_subquery_all)
            ).filter(
                is_deleted=False,
                branchstock_total_quantity__gt=0
            )

        return render(request, 'branchstock/branchstock_total.html', {'branches': Branch.objects.all(), 'products': products})
        
from .models import RequisitionBranchStock
from organization.models import Branch, EndDayRecord
from .forms import RequisitionBranchStockForm
class RequisitionBranchStockMixin(IsAdminMixin):
    model = RequisitionBranchStock
    form_class = RequisitionBranchStockForm
    paginate_by = 10
    queryset = RequisitionBranchStock.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('requisitionbranchstock_list')

    search_lookup_fields = [
        "product__title",
        "product__description",
    ]

class RequisitionBranchStockList(RequisitionBranchStockMixin, ListView):
    template_name = "requisitionbranchstock/requisitionbranchstock_list.html"

class RequisitionBranchStockDetail(RequisitionBranchStockMixin, DetailView):
    template_name = "requisitionbranchstock/requisitionbranchstock_detail.html"

class RequisitionBranchStockCreate(RequisitionBranchStockMixin, CreateView):
    template_name = "requisitionbranchstock/requisitionbranchstock_create.html"

class RequisitionBranchStockUpdate(RequisitionBranchStockMixin, UpdateView):
    template_name = "update.html"

class RequisitionBranchStockDelete(RequisitionBranchStockMixin, DeleteMixin, View):
    pass
