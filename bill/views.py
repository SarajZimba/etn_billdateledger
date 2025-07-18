from django.shortcuts import render
from django.contrib import messages

# Create your views here.

from django.db import transaction
from django.shortcuts import get_object_or_404, redirect

from django.http import HttpResponse, HttpResponseRedirect

from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from django.db.models import DateField
from django.db.models.functions import Cast
from django.db.models import Sum, Count, F, Q, Avg

import nepali_datetime
import pickle
from datetime import datetime, date


from root.utils import DeleteMixin
from user.permission import IsAccountantMixin, IsAdminMixin, BillFilterMixin, AdminBillingMixin
from .models import Bill, BillItem, TablReturnEntry
from .forms import BillForm, BillItemFormset
from .resources import (
    TblTaxEntryResource,
    SalesEntryResource,
    ReturnEntryResource,
)
from organization.models import Organization, Branch
from product.models import Product, BranchStock
from bill.models import BillItem
from user.models import Customer
import xlwt


class BranchMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["branches"] = Branch.objects.active()
        return context


class TransactionFilterMixin:
    search_lookup_fields = ["bill_no", "customer_name", "customer_pan"]

    def get_data(self, model):
        from_date = self.request.GET.get("fromDate", None)
        to_date = self.request.GET.get("toDate", None)
        query = self.request.GET.get("q", None)
        branch = self.request.GET.get("branch_code", None)
        if from_date and to_date:
            data = model.objects.annotate(
                bill_as_date=Cast("bill_date", DateField())
            ).filter(bill_as_date__range=(from_date, to_date))
        else:
            today_date = datetime.today().strftime("%Y-%m-%d")
            data = model.objects.annotate(
                bill_as_date=Cast("bill_date", DateField())
            ).filter(bill_as_date__range=(today_date, today_date))
        if query:
            q_lookup = Q()
            for field in self.search_lookup_fields:
                if field == "bill_no":
                    q_lookup |= Q(**{field + "__iexact": query})
                else:
                    q_lookup |= Q(**{field + "__icontains": query})
            data = data.filter(q_lookup)
        if branch:
            data = data.filter(bill_no__contains=branch)
        return data

from datetime import timedelta
class DateMixin:
    def get_date_data(self):
        from_date = self.request.GET.get("fromDate", None)
        to_date = self.request.GET.get("toDate", None)
        # # Convert to_date to a datetime object and add one day if it exists
        # if to_date:
        #     to_date = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        #     to_date = to_date.strftime('%Y-%m-%d')
        if to_date:
            to_date = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
            to_date = to_date.strftime('%Y-%m-%d')
        return from_date, to_date


class BillFilterDateMixin(DateMixin):
    search_lookup_fields = ["invoice_number", "customer_name", "customer_tax_number"]

    def get_queryset(self):
        qs = super().get_queryset()
        branch_code = self.request.GET.get("branch_code", None)
        query = self.request.GET.get("q", None)
        from_date, to_date = self.get_date_data()
        if from_date and to_date:
            qs = qs.filter(transaction_date__range=(from_date, to_date))
        else:
            today_date = datetime.today().strftime("%Y-%m-%d")
            qs = qs.filter(transaction_date__range=(today_date, today_date))

        if branch_code:
            qs = qs.filter(branch__branch_code=branch_code)

        if query:
            q_lookup = Q()
            for field in self.search_lookup_fields:
                if field == "invoice_number":
                    q_lookup |= Q(**{field + "__iexact": query})
                else:
                    q_lookup |= Q(**{field + "__icontains": query})
            qs = qs.filter(q_lookup)
        return qs


class ExportExcelMixin(BranchMixin):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["org"] = Organization.objects.last()
        return context

    def get(self, request, *args, **kwargs):
        self.format = request.GET.get("format", False)
        if self.format == "xls":
            return self.export_to_xls()
        return super().get(request, *args, **kwargs)

    def get_format_date(self, value):
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")

    def get_filter_date(self):
        fromDate = self.request.GET.get("fromDate", None)
        toDate = self.request.GET.get("toDate")
        if fromDate and toDate:
            return (self.get_format_date(fromDate), self.get_format_date(toDate))
        today = datetime.today().strftime("%d/%m/%Y")
        return (
            today,
            today,
        )

    def get_branch(self):
        return self.request.GET.get("branch_code", None)

    def init_xls(self, title, columns):
        organization = Organization.objects.last()
        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet(title)

        # Sheet header, first row
        row_num = 5

        font_style_organization = xlwt.XFStyle()
        font_style_organization = xlwt.easyxf(
            "font: bold 1,height 280; align:  horiz center"
        )
        font_style_organization.font.bold = True

        font_style_bold = xlwt.XFStyle()
        font_style_bold.font.bold = True

        font_style_normal = xlwt.XFStyle()

        ws.write_merge(
            0, 0, 0, len(columns) - 1, organization.org_name, font_style_organization
        )
        ws.row(0).height = 256 * 2
        ws.write_merge(
            1,
            1,
            0,
            len(columns) - 1,
            organization.company_address,
            font_style_organization,
        )
        ws.row(1).height = 256 * 2

        ws.write_merge(
            2,
            2,
            0,
            len(columns) - 1,
            title,
            font_style_organization,
        )
        ws.row(2).height = 256 * 2
        if self.get_filter_date():
            filter_date = (
                f"From: {self.get_filter_date()[0]} to {self.get_filter_date()[1]}"
            )
            ws.write_merge(
                3,
                3,
                0,
                len(columns) - 1,
                filter_date,
                font_style_organization,
            )
        ws.row(3).height = 256 * 2

        if self.get_branch():
            branch = f"Branch: {self.get_branch()}"
            ws.write_merge(
                4,
                4,
                0,
                len(columns) - 1,
                branch,
                font_style_organization,
            )

        ws.row(4).height = 256 * 2

        ws.write(row_num, 0, title, font_style_bold)
        row_num += 1

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style_bold)

        return wb, ws, row_num, font_style_normal, font_style_bold


class BillMixin(AdminBillingMixin):
    model = Bill
    form_class = BillForm
    paginate_by = 50
    queryset = Bill.objects.filter(is_deleted=False).order_by('-created_at')
    success_url = reverse_lazy("bill_list")
    search_lookup_fields = [
        "customer_name",
        "invoice_number",
        "customer_tax_number",
        "terminal",
    ]

    def get_queryset(self, *args, **kwargs):
        qc = super().get_queryset(*args, **kwargs)
        qc = self.search(qc)
        qc = self.date_filter(qc)
        qc = self.terminalSearch(qc)
        return qc

    def date_filter(self, qc):
        if self.request.GET.get("sort_date"):
            sort_date = self.request.GET.get("sort_date")
            if sort_date in self.valid_date_filter:
                return qc.order_by(sort_date)
        return qc

    def search(self, qc):
        if self.request.GET.get("q"):
            query = self.request.GET.get("q")
            q_lookup = Q()
            for field in self.search_lookup_fields:
                q_lookup |= Q(**{field + "__iexact": query})
            return qc.filter(q_lookup)

        return qc

    def terminalSearch(self, qc):

        if self.request.GET.get("terminal"):
            query = self.request.GET.get("terminal")
            q_lookup = Q()
            for field in self.search_lookup_fields:
                q_lookup |= Q(**{field + "__icontains": query})
            return qc.filter(q_lookup)

        return qc


class SalesEntryMixin(TransactionFilterMixin):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.filter_by_queryset:
            context["sales_entry"] = self.get_queryset()
        else:
            context["sales_entry"] = self.get_data(model=TblSalesEntry)

        context["total_sales_amount"] = context["sales_entry"].aggregate(Sum("amount"))[
            "amount__sum"
        ]

        context["total_no_tax_sales"] = context["sales_entry"].aggregate(
            Sum("NoTaxSales")
        )["NoTaxSales__sum"]

        context["total_tax_amount"] = context["sales_entry"].aggregate(
            Sum("tax_amount")
        )["tax_amount__sum"]

        context["export_at_zero_rate"] = context["sales_entry"].aggregate(
            Sum("ZeroTaxSales")
        )["ZeroTaxSales__sum"]

        context["total_taxable_amount"] = context["sales_entry"].aggregate(
            Sum("taxable_amount")
        )["taxable_amount__sum"]

        context["exempted_sales"] = context["sales_entry"].aggregate(
            Sum("exemptedSales")
        )["exemptedSales__sum"]

        return context


class ReturnEntryMixin(TransactionFilterMixin):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.filter_by_queryset:
            context["return_entry"] = self.get_queryset()
        else:
            context["return_entry"] = self.get_data(model=TablReturnEntry)

        context["total_return_amount"] = (
            context["return_entry"]
            .aggregate(amount_sum=Sum("amount"))
            .get("amount_sum")
            or 0
        )
        context["total_return_tax_amount"] = (
            context["return_entry"].aggregate(Sum("tax_amount"))["tax_amount__sum"] or 0
        )
        context["total_no_tax_return"] = (
            context["return_entry"].aggregate(Sum("NoTaxSales"))["NoTaxSales__sum"] or 0
        )
        context["export_at_zero_rate_return"] = (
            context["return_entry"].aggregate(Sum("ZeroTaxSales"))["ZeroTaxSales__sum"]
            or 0
        )
        context["total_taxable_return_amount"] = (
            context["return_entry"].aggregate(Sum("taxable_amount"))[
                "taxable_amount__sum"
            ]
            or 0
        )

        context["return_exempted_sales"] = (
            context["return_entry"].aggregate(Sum("exemptedSales"))[
                "exemptedSales__sum"
            ]
            or 0
        )
        return context


class SalesSummaryMixin(SalesEntryMixin, ReturnEntryMixin):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["org"] = Organization.objects.last()
        if context["return_entry"] or context["sales_entry"]:
            context["grand_total_amount"] = self.get_difference(
                context["total_sales_amount"], context["total_return_amount"]
            )
            context["grand_total_no_tax"] = self.get_difference(
                context["total_no_tax_sales"], context["total_no_tax_return"]
            )
            context["grand_total_export_at_zero_rate"] = self.get_difference(
                context["export_at_zero_rate"], context["export_at_zero_rate_return"]
            )
            context["grand_total_taxable_amount"] = self.get_difference(
                context["total_taxable_amount"], context["total_taxable_return_amount"]
            )

            context["grand_total_tax_amount"] = self.get_difference(
                context["total_tax_amount"], context["total_return_tax_amount"]
            )

            context["grand_total_exempted_sales"] = self.get_difference(
                context["exempted_sales"], context["return_exempted_sales"]
            )

        return context


class IncrementBillPrintCount(BillMixin, View):
    def get(self, request, *args, **kwargs):
        bill = get_object_or_404(Bill, pk=self.request.GET.get("pk"))
        bill.print_count += 1
        bill.save()
        return HttpResponse("success")


class MarkBillVoid(BillMixin, View):
    def post(self, request, *args, **kwargs):
        reason = request.POST.get("voidReason")
        bill = get_object_or_404(Bill, pk=self.kwargs.get("id"))
        bill.status = False
        bill.save()

        miti = bill.transaction_miti
        quantity = bill.bill_items.count()
        return_entry = TablReturnEntry(
            bill_date=bill.transaction_date,
            bill_no=bill.invoice_number,
            customer_name=bill.customer_name,
            customer_pan=bill.customer_tax_number,
            amount=bill.grand_total,
            NoTaxSales=0,
            ZeroTaxSales=0,
            taxable_amount=bill.taxable_amount,
            tax_amount=bill.tax_amount,
            miti=miti,
            ServicedItem="Goods",
            quantity=quantity,
            reason=reason,
            excise_duty_amount=bill.excise_duty_amount
        )
        return_entry.save()
        try:
            entry = TblTaxEntry.objects.get(bill_no=bill.invoice_number)
            entry.is_active="NO"
            entry.save()
        except Exception:
            pass

        return redirect(
            reverse_lazy("bill_detail", kwargs={"pk": self.kwargs.get("id")})
        )


class BillList(BillMixin, ListView):
    template_name = "bill/bill_list.html"
    def get_queryset(self):
        queryset = super().get_queryset()

        from_date = self.request.GET.get('fromDate')
        to_date = self.request.GET.get('toDate')

        if from_date and to_date:
            # Convert the string representations of dates to datetime objects
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d')

            # Add one day to the 'to_date'
            one_day = timedelta(days=1)
            to_date_obj += one_day

            # Convert the datetime objects back to strings in the same format
            from_date = from_date_obj.strftime('%Y-%m-%d')
            to_date = to_date_obj.strftime('%Y-%m-%d')
            print(to_date)

            # Filter the queryset using the updated date range
            queryset = queryset.filter(created_at__range=[from_date, to_date])

        return queryset



class SalesInvoiceSummaryRegister(
    BillFilterDateMixin, BillMixin, ExportExcelMixin, ListView
):
    template_name = "bill/sales_invoice_summary.html"
    queryset = Bill.objects.filter(is_deleted=False, status=True)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        if queryset:
            context["total_sales_amount"] = queryset.aggregate(Sum("sub_total"))[
                "sub_total__sum"
            ]
            context["total_discount_amount"] = queryset.aggregate(
                Sum("discount_amount")
            )["discount_amount__sum"]
            context["total_tax_amount"] = queryset.aggregate(Sum("tax_amount"))[
                "tax_amount__sum"
            ]
            context["total_grand_total"] = queryset.aggregate(Sum("grand_total"))[
                "grand_total__sum"
            ]
        return context

    def export_to_xls(self):
        response = HttpResponse(content_type="application/ms-excel")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="sales_invoice_summary.xls"'

        columns = [
            "transaction_date",
            "invoice_number",
            "customer",
            "sub_total",
            "discount_amount",
            "tax_amount",
            "grand_total",
            "product_discount",
            "other_terms",
        ]
        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Sales Invoice Summary", columns
        )

        rows = self.get_queryset().values_list(*columns[:7])
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style_normal)
            ws.write(row_num, 7, "0", font_style_normal)
            ws.write(row_num, 8, "0", font_style_normal)

        sum_sales_invoice = self.get_queryset().aggregate(
            sub_total=Sum(
                "sub_total",
            ),
            discount_amount=Sum(
                "discount_amount",
            ),
            tax_amount=Sum(
                "tax_amount",
            ),
            grand_total=Sum(
                "grand_total",
            ),
        )

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)

        for key, value in sum_sales_invoice.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        wb.save(response)
        return response


class BillDetail(BillMixin, DetailView):
    template_name = "bill/bill_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tbl_return = TablReturnEntry.objects.filter(
            bill_no=self.object.invoice_number
        ).first()
        if tbl_return:
            context["reason"] = tbl_return.reason
        return context


class BillCreate(BillMixin, CreateView):
    template_name = "bill/bill_create.html"

    def get_success_url(self):
        messages.success(self.request, "Bill Created Successfully")
        return reverse_lazy("bill_detail", kwargs={"pk": self.object.id})

    def form_valid(self, form):
        nepali_today = nepali_datetime.date.today()
        print("form_instance", form.instance)
        print("form_instance_excise_duty", form.instance.excise_duty_amount)
        excise_duty_amount = form.instance.excise_duty_amount
        form.instance.organization = self.request.user.organization
        if Branch.objects.active().filter(is_central_billing=True).last():
            form.instance.branch = Branch.objects.active().filter(is_central_billing=True).last()
        else:
            form.instance.branch = Branch.objects.active().last()
        form.instance.transaction_miti = nepali_today
        form.instance.agent = self.request.user
        form.instance.agent_name = self.request.user.full_name
        form.instance.terminal = 1
        if form.instance.payment_mode.lower() == 'credit' and not form.instance.customer:
            print(form.instance.payment_mode.lower())
            print(form.instance.customer)
            # return redirect('bill_create')
            messages.error(self.request, "Please  fill in the customer form.")
            return self.form_invalid(form)
        
        self.object = form.save()
        context = self.get_context_data()
        self.save_bill_item_attributes()
        return super().form_valid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)

    def save_bill_item_attributes(self):
        req = self.request.POST
        if Branch.objects.active().filter(is_central_billing=True).last():
            branch = Branch.objects.active().filter(is_central_billing=True).last()
        else:
            branch = Branch.objects.active().last()
        products = req.getlist("product")
        discount = req.get("discount_amount", 0)
        sub_total = 0
        for id in products:
            product = Product.objects.get(id=id)
            rate = req.get(f"id_bill_item_rate_{id}")
            product_quantity = req.get(f"id_bill_item_quantity_{id}")
            quantity = int(product_quantity)
            is_taxable = req.get("is_taxable")
            amount = float(rate) * float(product_quantity)
            sub_total += amount
            bill_item = BillItem.objects.create(
                product_title=product.title,
                product_quantity=product_quantity,
                unit_title=product.unit,
                rate=rate,
                amount=amount,
                product=product,
                agent=self.request.user,
            )
            org = Organization.objects.first()
            if org.allow_negative_sales == False:
                branchstock_entries = BranchStock.objects.filter(branch=branch, product=product, is_deleted=False).order_by('quantity')
                
                for branchstock in branchstock_entries:
                        if quantity > 0:
                            available_quantity = branchstock.quantity
                            if quantity >= available_quantity:
                                    # Reduce the available quantity to 0 and update the branchstock entry
                                branchstock.quantity = 0
                                branchstock.save()
                                quantity -= available_quantity
                            else:
                                    # Reduce the available quantity by the bill item quantity and update the branchstock entry
                                branchstock.quantity -= quantity
                                branchstock.save()
                                quantity = 0
            self.object.bill_items.add(bill_item)
        self.object.sub_total = round(sub_total, 2)


class BillUpdate(BillMixin, UpdateView):
    template_name = "update.html"


class BillDelete(BillMixin, DeleteMixin, View):
    pass


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import TblTaxEntry
from .forms import TblTaxEntryForm


# this is materialize view
class TblTaxEntryMixin(AdminBillingMixin):
    model = TblTaxEntry
    form_class = TblTaxEntryForm
    paginate_by = 50
    queryset = TblTaxEntry.objects.all()
    success_url = reverse_lazy("tbltaxentry_list")
    search_lookup_fields = [
        "customer_name",
        "customer_pan",
        "bill_no",
    ]

    def get_queryset(self, *args, **kwargs):
        qc = super().get_queryset(*args, **kwargs).order_by("-idtbltaxEntry")
        return qc


class TblTaxEntryList(TblTaxEntryMixin, ListView):
    template_name = "tbltaxentry/tbltaxentry_list.html"
    queryset = TblTaxEntry.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tax_entry = self.get_queryset()
        context["total_amount"] = tax_entry.aggregate(Sum("amount"))["amount__sum"]
        context["discount"] = tax_entry.aggregate(Sum("discount"))["discount__sum"] or 0
        context["taxable_amount"] = tax_entry.aggregate(Sum("taxable_amount"))[
            "taxable_amount__sum"
        ]
        context["tax_amount"] = tax_entry.aggregate(Sum("tax_amount"))[
            "tax_amount__sum"
        ]
        context["vat_refund_amount"] = tax_entry.aggregate(Sum("vat_refund_amount"))[
            "vat_refund_amount__sum"
        ]

        return context


class TblTaxEntryDetail(TblTaxEntryMixin, DetailView):
    template_name = "tbltaxentry/tbltaxentry_detail.html"


class TblTaxEntryCreate(TblTaxEntryMixin, CreateView):
    template_name = "create.html"


class TblTaxEntryUpdate(TblTaxEntryMixin, UpdateView):
    template_name = "update.html"


class TblTaxEntryDelete(TblTaxEntryMixin, DeleteMixin, View):
    pass


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import TblSalesEntry
from .forms import TblSalesEntryForm

from user.permission import PartyAdminBillingMixin
# class TblSalesEntryMixin(AdminBillingMixin):
class TblSalesEntryMixin(PartyAdminBillingMixin):
    model = TblSalesEntry
    form_class = TblSalesEntryForm
    paginate_by = 50
    success_url = reverse_lazy("tblsalesentry_list")
    search_lookup_fields = ["customer_name", "customer_pan"]


class TblSalesEntryList(BranchMixin, SalesEntryMixin, TblSalesEntryMixin, ListView):
    template_name = "tblsalesentry/tblsalesentry_list.html"
    queryset = TblSalesEntry.objects.all()
    filter_by_queryset = False

    def get_queryset(self, *args, **kwargs):
        qc = super().get_queryset(*args, **kwargs).order_by("-tblSalesEntry")
        return qc


class TblSalesEntryDetail(TblSalesEntryMixin, DetailView):
    template_name = "tblsalesentry/tblsalesentry_detail.html"


class TblSalesEntryCreate(TblSalesEntryMixin, CreateView):
    template_name = "create.html"


class TblSalesEntryUpdate(TblSalesEntryMixin, UpdateView):
    template_name = "update.html"


class TblSalesEntryDelete(TblSalesEntryMixin, DeleteMixin, View):
    pass


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import TablReturnEntry
from .forms import TablReturnEntryForm


class TablReturnEntryMixin(AdminBillingMixin):
    model = TablReturnEntry
    form_class = TablReturnEntryForm
    paginate_by = 50
    queryset = TablReturnEntry.objects.all()
    success_url = reverse_lazy("tablreturnentry_list")

    def get_queryset(self, *args, **kwargs):
        qc = super().get_queryset(*args, **kwargs).order_by("-idtblreturnEntry")
        return qc


class TablReturnEntryList(
    ExportExcelMixin, ReturnEntryMixin, TablReturnEntryMixin, ListView
):
    template_name = "tablreturnentry/tablreturnentry_list.html"
    queryset = TablReturnEntry.objects.all()
    filter_by_queryset = False

    def export_to_xls(self):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="sales_return.xls"'
        columns = [
            "miti",
            "bill_no",
            "customer_name",
            "customer_pan",
            "amount",
            "taxable_amount",
            "tax_amount",
            "reason",
        ]
        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Sales Return", columns
        )

        sales_return = self.get_data(TablReturnEntry)

        rows = sales_return.values_list(*columns)
        sum_return_entry = sales_return.aggregate(
            amount=Sum(
                "amount",
            ),
            taxable_amount=Sum(
                "taxable_amount",
            ),
            tax_amount=Sum(
                "tax_amount",
            ),
        )

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if columns[col_num] == "miti":
                    new_val = ""
                    if row[col_num]:
                        new_val = (row[col_num].replace("-", "."),)
                    ws.write(
                        row_num,
                        col_num,
                        new_val,
                        font_style_normal,
                    )
                else:
                    ws.write(row_num, col_num, row[col_num], font_style_normal)
        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)
        for key, value in sum_return_entry.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        wb.save(response)
        return response


class TablReturnEntryDetail(TablReturnEntryMixin, DetailView):
    template_name = "tablreturnentry/tablreturnentry_detail.html"


class TablReturnEntryCreate(TablReturnEntryMixin, CreateView):
    template_name = "create.html"


class TablReturnEntryUpdate(TablReturnEntryMixin, UpdateView):
    template_name = "update.html"


class TablReturnEntryDelete(TablReturnEntryMixin, DeleteMixin, View):
    pass


# this is for sales book
class MaterializedView(
    ExportExcelMixin, SalesSummaryMixin, AdminBillingMixin, TemplateView
):
    template_name = "bill/materialized_view.html"
    filter_by_queryset = False

    def get_difference(self, data1, data2):
        if isinstance(data1, type(None)):
            return -data2
        if isinstance(data2, type(None)):
            return data1
        if data1 >= data2:
            return data1 - data2
        return 0

    def date_range_filter(self, qc):
        if self.request.GET.get("fromDate") and self.request.GET.get("toDate"):
            created_at_date = self.request.GET.get("fromDate")
            created_to_date = self.request.GET.get("toDate")
            return qc.filter(bill_date__range=[created_at_date, created_to_date])
        return qc

    def export_to_xls(self):
        organization = Organization.objects.last()
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="sales_book.xls"'

        common = [
            "miti",
            "bill_date",
            "bill_no",
            "customer_name",
            "customer_pan",
            "amount",
            "ServicedItem",
            "quantity",
            # "unit",
            "NoTaxSales",
            "ZeroTaxSales",
            "taxable_amount",
            "tax_amount",
            "exemptedSales",
            "export",
            "exportCountry",
            "exportNumber",
            "exportDate",
        ]
        columns = [
            "tblSalesEntry",
        ] + common

        new_columns = ["id"] + columns[1:]

        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Sales Book", new_columns
        )

        sales_entry = self.get_data(TblSalesEntry)
        rows = sales_entry.values_list(*columns)
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if columns[col_num] == "miti":
                    new_val = ""
                    if row[col_num]:
                        new_val = (row[col_num].replace("-", "."),)
                    ws.write(
                        row_num,
                        col_num,
                        new_val,
                        font_style_normal,
                    )
                else:
                    ws.write(row_num, col_num, row[col_num], font_style_normal)

        sum_sales_entry = sales_entry.aggregate(
            amount=Sum(
                "amount",
            ),
            NoTaxSales=Sum(
                "NoTaxSales",
            ),
            ZeroTaxSales=Sum(
                "ZeroTaxSales",
            ),
            taxable_amount=Sum(
                "taxable_amount",
            ),
            tax_amount=Sum(
                "tax_amount",
            ),
            exemptedSales=Sum(
                "exemptedSales",
            ),
        )

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)
        for key, value in sum_sales_entry.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        columns2 = [
            "idtblreturnEntry",
        ] + common

        row_num += 1
        ws.write(row_num, 0, "")
        row_num += 1
        ws.write(row_num, 0, "Sales Return", font_style_bold)
        row_num += 1

        new_columns = ["id"] + columns2[1:]
        for col_num in range(len(columns2)):
            ws.write(row_num, col_num, new_columns[col_num], font_style_bold)

        sales_return = self.get_data(TablReturnEntry)
        rows2 = sales_return.values_list(*columns2)
        sum_return_entry = sales_return.aggregate(
            amount=Sum(
                "amount",
            ),
            NoTaxSales=Sum(
                "NoTaxSales",
            ),
            ZeroTaxSales=Sum(
                "ZeroTaxSales",
            ),
            taxable_amount=Sum(
                "taxable_amount",
            ),
            tax_amount=Sum(
                "tax_amount",
            ),
            exemptedSales=Sum(
                "exemptedSales",
            ),
        )

        for row in rows2:
            row_num += 1
            for col_num in range(len(row)):
                if columns[col_num] == "miti":
                    new_val = ""
                    if row[col_num]:
                        new_val = (row[col_num].replace("-", "."),)
                    ws.write(
                        row_num,
                        col_num,
                        new_val,
                        font_style_normal,
                    )
                else:
                    ws.write(row_num, col_num, row[col_num], font_style_normal)
        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)
        for key, value in sum_return_entry.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        row_num += 2
        ws.write(row_num, 0, "Grand Total", font_style_bold)
        grand_total = {
            k: self.get_difference(sum_sales_entry[k], sum_return_entry[k])
            for k in sum_sales_entry
        }
        for key, value in grand_total.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_bold)
        wb.save(response)
        return response


class MaterializedViewExportExcel(View):
    def get(self, request, *args, **kwargs):
        organization = Organization.objects.last()
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="materialized_view.xls"'

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Materialized View")

        # Sheet header, first row
        row_num = 4

        font_style_organization = xlwt.XFStyle()
        font_style_organization = xlwt.easyxf(
            "font: bold 1,height 280; align:  horiz center"
        )
        font_style_organization.font.bold = True

        font_style_bold = xlwt.XFStyle()
        font_style_bold.font.bold = True

        font_style_normal = xlwt.XFStyle()

        font_style_date = xlwt.XFStyle()
        font_style_date.num_format_str = "yyyy.mm.dd"

        columns = [
            "idtbltaxEntry",
            "fiscal_year",
            "bill_no",
            "customer_name",
            "customer_pan",
            "bill_date",
            "amount",
            "discount",
            "taxable_amount",
            "tax_amount",
            "is_printed",
            "is_active",
            "printed_time",
            "entered_by",
            "printed_by",
            "is_realtime",
            "sync_with_ird",
            "payment_method",
            "vat_refund_amount",
            "transaction_id",
        ]
        ws.write_merge(
            0, 0, 0, len(columns) - 1, organization.org_name, font_style_organization
        )
        ws.row(0).height = 256 * 2
        ws.write_merge(
            1,
            1,
            0,
            len(columns) - 1,
            organization.company_address,
            font_style_organization,
        )
        ws.row(1).height = 256 * 2

        ws.write_merge(
            2,
            2,
            0,
            len(columns) - 1,
            "Materialized view",
            font_style_organization,
        )
        ws.row(1).height = 256 * 2

        new_columns = ["id"] + columns[1:]
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, new_columns[col_num], font_style_bold)

        rows = TblTaxEntry.objects.all().values_list(*columns)
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if columns[col_num] == "bill_date":
                    ws.write(
                        row_num,
                        col_num,
                        row[col_num],
                        font_style_date,
                    )
                else:
                    ws.write(row_num, col_num, row[col_num], font_style_normal)

        wb.save(response)
        return response


class SalesEntryViewExportExcel(TransactionFilterMixin, View):
    def get(self, request, *args, **kwargs):
        organization = Organization.objects.last()
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="sales_book.xls"'

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Sales Entry and Bill Return")

        # Sheet header, first row
        row_num = 4

        font_style_organization = xlwt.XFStyle()
        font_style_organization = xlwt.easyxf(
            "font: bold 1,height 280; align:  horiz center"
        )
        font_style_organization.font.bold = True

        font_style_bold = xlwt.XFStyle()
        font_style_bold.font.bold = True

        font_style_normal = xlwt.XFStyle()

        common = [
            "miti",
            # "bill_date",
            "bill_no",
            "customer_name",
            "customer_pan",
            "amount",
            "ServicedItem",
            "quantity",
            # "unit",
            "NoTaxSales",
            "ZeroTaxSales",
            "taxable_amount",
            "tax_amount",
            "exemptedSales",
            "export",
            "exportCountry",
            "exportNumber",
            "exportDate",
        ]
        columns = [
            "tblSalesEntry",
        ] + common

        ws.write_merge(
            0, 0, 0, len(columns) - 1, organization.org_name, font_style_organization
        )
        ws.row(0).height = 256 * 2
        ws.write_merge(
            1,
            1,
            0,
            len(columns) - 1,
            organization.company_address,
            font_style_organization,
        )
        ws.row(1).height = 256 * 2

        ws.write_merge(
            2,
            2,
            0,
            len(columns) - 1,
            "Sales Book",
            font_style_organization,
        )
        ws.row(2).height = 256 * 2

        ws.write(row_num, 0, "Sales Entry", font_style_bold)
        row_num += 1

        new_columns = ["id"] + columns[1:]
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, new_columns[col_num], font_style_bold)

        sales_entry = self.get_data(TblSalesEntry)
        rows = sales_entry.values_list(*columns)
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if columns[col_num] == "miti":
                    new_val = ""
                    if row[col_num]:
                        new_val = (row[col_num].replace("-", "."),)
                    ws.write(
                        row_num,
                        col_num,
                        new_val,
                        font_style_normal,
                    )
                else:
                    ws.write(row_num, col_num, row[col_num], font_style_normal)

        sum_sales_entry = sales_entry.aggregate(
            amount=Sum(
                "amount",
            ),
            NoTaxSales=Sum(
                "NoTaxSales",
            ),
            ZeroTaxSales=Sum(
                "ZeroTaxSales",
            ),
            taxable_amount=Sum(
                "taxable_amount",
            ),
            tax_amount=Sum(
                "tax_amount",
            ),
            exemptedSales=Sum(
                "exemptedSales",
            ),
        )

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)
        for key, value in sum_sales_entry.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        columns2 = [
            "idtblreturnEntry",
        ] + common

        row_num += 1
        ws.write(row_num, 0, "")
        row_num += 1
        ws.write(row_num, 0, "Sales Return", font_style_bold)
        row_num += 1

        new_columns = ["id"] + columns2[1:]
        for col_num in range(len(columns2)):
            ws.write(row_num, col_num, new_columns[col_num], font_style_bold)

        sales_return = self.get_data(TablReturnEntry)
        rows2 = sales_return.values_list(*columns2)
        sum_return_entry = sales_return.aggregate(
            amount=Sum(
                "amount",
            ),
            NoTaxSales=Sum(
                "NoTaxSales",
            ),
            ZeroTaxSales=Sum(
                "ZeroTaxSales",
            ),
            taxable_amount=Sum(
                "taxable_amount",
            ),
            tax_amount=Sum(
                "tax_amount",
            ),
            exemptedSales=Sum(
                "exemptedSales",
            ),
        )

        for row in rows2:
            row_num += 1
            for col_num in range(len(row)):
                if columns[col_num] == "miti":
                    new_val = ""
                    if row[col_num]:
                        new_val = (row[col_num].replace("-", "."),)
                    ws.write(
                        row_num,
                        col_num,
                        new_val,
                        font_style_normal,
                    )
                else:
                    ws.write(row_num, col_num, row[col_num], font_style_normal)
        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)
        for key, value in sum_return_entry.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        row_num += 2
        # ws.write(row_num, 0, "Grand Total", font_style_bold)
        # grand_total = {
        #     k: sum_sales_entry[k] - sum_return_entry[k] for k in sum_sales_entry
        # }
        # for key, value in grand_total.items():
        #     ws.write(row_num, columns.index(key), value or 0, font_style_bold)

        wb.save(response)
        return response


class ReturnEntryViewExportExcel(View):
    def get(self, request, *args, **kwargs):
        dataset = ReturnEntryResource().export()
        response = HttpResponse(dataset.xls, content_type="application/vnd.ms-excel")
        response["Content-Disposition"] = 'attachment; filename="exported_data.xls"'
        return response


class CustomerWiseSalesInvoiceRegister(BillFilterDateMixin, ExportExcelMixin, ListView):
    template_name = "bill/report/customer_wise_sales.html"

    queryset = Bill.objects.values("customer__name").annotate(
        quantity=Sum("bill_items__product_quantity"),
        vat=Sum("tax_amount"),
        amount=Sum(F("bill_items__product_quantity") * F("bill_items__rate")),
    )

    def export_to_xls(self):
        response = HttpResponse(content_type="application/ms-excel")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="sales_invoice_register.xls"'
        columns = ["customer__name", "quantity", "amount", "vat", "net_amount"]
        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Sales Invoice Register", columns
        )

        rows = self.get_queryset().values_list(*columns[:4])
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style_normal)
            ws.write(row_num, 4, row[2] + row[3], font_style_normal)

        total = self.get_queryset().aggregate(
            quantity=Count("bill_items"),
            amount=Sum("grand_total"),
            vat=Sum("tax_amount"),
            net_amount=Sum(F("grand_total") + F("tax_amount")),
        )

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)

        for key, value in total.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        wb.save(response)
        return response


class PartyWiseSalesInvoiceRegister(
    DateMixin, TblSalesEntryMixin, ExportExcelMixin, ListView
):
    template_name = "bill/report/party_wise_sales.html"
    queryset = TblSalesEntry.objects.values_list(
        "customer_name", "customer_pan"
    ).annotate(
        amount=Sum("amount"),
        exempted_sales=Sum("exemptedSales"),
        taxable_amount=Sum("taxable_amount"),
        tax_amount=Sum("tax_amount"),
    )

    def get_filtered_date(self, qs):
        from_date, to_date = self.get_date_data()
        if from_date and to_date:
            qs = qs.filter(bill_date__range=(from_date, to_date))
        else:
            today_date = datetime.today().strftime("%Y-%m-%d")
            qs = qs.filter(bill_date__range=(today_date, today_date))
        return qs

    def get_queryset(self):
        qs = super().get_queryset()
        branch_code = self.request.GET.get("branch_code", None)
        qs = self.get_filtered_date(qs)
        if branch_code:
            qs = qs.filter(bill_no__contains=branch_code)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        qs = TblSalesEntry.objects.values_list("customer_name", "customer_pan")
        qs = self.get_filtered_date(qs)
        total = qs.aggregate(
            amount=Sum("amount"),
            exemptedSales=Sum("exemptedSales"),
            taxable_amount=Sum("taxable_amount"),
            tax_amount=Sum("tax_amount"),
        )

        context["total"] = total
        return context

    def export_to_xls(self):
        response = HttpResponse(content_type="application/ms-excel")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="partywise_sales_register.xls"'

        columns = [
            "customer_name",
            "customer_pan",
            "amount",
            "exemptedSales",
            "taxable_amount",
            "tax_amount",
            "export",
        ]
        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Party Wise Sales Register", columns
        )

        rows = self.get_queryset()
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style_normal)
            ws.write(row_num, 6, "0", font_style_normal)

        qs = TblSalesEntry.objects.values_list("customer_name", "customer_pan")
        qs = self.get_filtered_date(qs)
        total = qs.aggregate(
            amount=Sum("amount"),
            exemptedSales=Sum("exemptedSales"),
            taxable_amount=Sum("taxable_amount"),
            tax_amount=Sum("tax_amount"),
        )

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)

        for key, value in total.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        wb.save(response)
        return response


class SalesInvoiceAnalysis(BillFilterDateMixin, ExportExcelMixin, ListView):
    template_name = "bill/report/sales_invoice_analysis.html"
    queryset = Bill.objects.values("customer__id").annotate(
        quantity=Sum("bill_items__product_quantity"),
        rate=Avg("bill_items__rate"),
        net_amount=Sum(F("bill_items__product_quantity") * F("bill_items__rate")),
    )

    def get_sales_data(self):
        item = []
        from_date, to_date = self.get_date_data()
        today_date = datetime.today().strftime("%Y-%m-%d")
        for i in self.get_queryset():
            qs = BillItem.objects.filter(bill__customer_id=i["customer__id"])
            if from_date and to_date:
                qs = qs.filter(bill__transaction_date__range=(from_date, to_date))
            else:
                qs = qs.filter(bill__transaction_date__range=(today_date, today_date))
            qs = qs.values_list("product_title", "unit_title")
            reloaded_qs = BillItem.objects.active()
            reloaded_qs.query = pickle.loads(pickle.dumps(qs.query))
            bills = reloaded_qs.annotate(
                quantity=Sum("product_quantity"),
                rate_new=Avg("rate"),
                net_amount=Sum(F("product_quantity") * F("rate")),
            )
            item.append(
                {
                    "bills": bills,
                    "total": i,
                    "customer": Customer.objects.get(id=int(i["customer__id"]))
                    if i["customer__id"]
                    else None,
                }
            )
        return item

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["data"] = self.get_sales_data()
        return context

    def export_to_xls(self):
        response = HttpResponse(content_type="application/ms-excel")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="sales_invoice_analysis.xls"'

        columns = [
            "name",
            "unit",
            "quantity",
            "rate",
            "net_amount",
            "complimentary",
        ]
        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Sales Invoice Analysis", columns
        )

        rows = self.get_sales_data()
        for row in rows:
            if row["customer"]:
                row_num += 1
                ws.write(row_num, 0, row["customer"].name, font_style_bold)
                rows2 = row["bills"]
                for row2 in rows2:
                    row_num += 1
                    for index, value in enumerate(row2.values()):
                        if isinstance(value, float) or isinstance(value, int):
                            ws.write(row_num, index, round(value, 2), font_style_normal)
                        else:
                            ws.write(row_num, index, value, font_style_normal)
                    ws.write(row_num, 5, 0, font_style_normal)

                row_num += 1
                ws.write(row_num, 0, "Total", font_style_bold)
                for key, value in row["total"].items():
                    if key in columns:
                        ws.write(
                            row_num,
                            columns.index(key),
                            round(value, 2) if value else 0,
                            font_style_bold,
                        )

                row_num += 1

        wb.save(response)
        return response


class PaymentModeSummary(BillFilterDateMixin, ExportExcelMixin, ListView):
    template_name = "bill/report/payment_mode_summary.html"
    queryset = Bill.objects.active().annotate(
        quantity=Sum("bill_items__product_quantity"),
    )

    def get_queryset(self):
        qs = super().get_queryset()
        payment_mode = self.request.GET.get("payment_mode", None)
        if payment_mode:
            qs = qs.filter(payment_mode=payment_mode)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["total"] = self.get_queryset().aggregate(
            sub_total=Sum("sub_total"),
            discount_amount=Sum("discount_amount"),
            tax_amount=Sum("tax_amount"),
            grand_total=Sum("grand_total"),
        )

        return context

    def export_to_xls(self):
        response = HttpResponse(content_type="application/ms-excel")
        response[
            "Content-Disposition"
        ] = 'attachment; filename="payment_mode_summary.xls"'
        columns = [
            "transaction_date",
            "invoice_number",
            "customer__name",
            "sub_total",
            "discount_amount",
            "tax_amount",
            "grand_total",
            "payment_mode",
        ]
        readable_columns_name = [
            "Date",
            "Invoice Number",
            "Customer Name",
            "Net Amount",
            "Discount",
            "VAT",
            "Total",
            "Payment Mode",
        ]
        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Bill Summary Details", readable_columns_name
        )

        rows = self.get_queryset().values_list(*columns)
        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                value = row[col_num]
                if isinstance(value, date):
                    ws.write(
                        row_num,
                        col_num,
                        str(value.strftime("%d/%m/%Y")),
                        font_style_normal,
                    )
                else:
                    ws.write(row_num, col_num, value, font_style_normal)

        total = self.get_queryset().aggregate(
            sub_total=Sum("sub_total"),
            discount_amount=Sum("discount_amount"),
            tax_amount=Sum("tax_amount"),
            grand_total=Sum("grand_total"),
        )

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)

        for key, value in total.items():
            ws.write(row_num, columns.index(key), value or 0, font_style_normal)

        wb.save(response)
        return response


from decimal import Decimal
class TodaysTransactionView(View):

    def get(self, request):
        today = date.today()
        bills = Bill.objects.filter(transaction_date=today, status=True)
        last_update = Bill.objects.order_by('-created_at').first().created_at if Bill.objects.order_by('-created_at') else None
        
        terminals = {}
        for bill in bills:
            if bill.invoice_number:
                bill_no_lst = bill.invoice_number.split('-')[:2]
                terminal_no = f"{bill_no_lst[0]}-{bill_no_lst[1]}"
                if terminal_no not in terminals:
                    last_updated = Bill.objects.filter(invoice_number__startswith=terminal_no).order_by('-created_at').first().created_at
                    terminals[terminal_no] = {"total_sale":0, "vat":0,
                                            "net_sale":0, "discount":0,
                                            "cash":0,"food":0, "beverage":0, "others":0,"credit_card":0, "mobile_payment":0, 'last_updated':last_updated }
                
                terminals[terminal_no]['total_sale'] += bill.grand_total
                terminals[terminal_no]['vat'] += bill.tax_amount
                terminals[terminal_no]['discount'] += bill.discount_amount
                terminals[terminal_no]['net_sale'] += (bill.grand_total-bill.tax_amount)

                if bill.payment_mode.lower().strip() == "cash":
                    terminals[terminal_no]['cash'] += bill.grand_total
                elif bill.payment_mode.lower().strip() == "credit card":
                    terminals[terminal_no]['credit_card'] += bill.grand_total
                elif bill.payment_mode.lower().strip() == "mobile payment":
                    terminals[terminal_no]['mobile_payment'] += bill.grand_total
                
                for item in bill.bill_items.all():
                    if item.product.category.title.lower().strip() == "food":
                        terminals[terminal_no]['food'] += item.amount
                    elif item.product.category.title.lower().strip() == "beverage":
                        terminals[terminal_no]['beverage'] += item.amount
                    elif item.product.category.title.lower().strip() == "others":
                        terminals[terminal_no]['others'] += item.amount
                

        terminals_to_template = []
        for k,v in terminals.items():
            v['terminal'] = k
            terminals_to_template.append(v)

        print(terminals_to_template)
        terminal_total_sale = Decimal(0.0)
        terminal_net_sale = Decimal(0.0)
        terminal_vat = Decimal(0.0)
        terminal_discount = Decimal(0.0)
        terminal_credit_card = Decimal(0.0)
        terminal_mobile_payment = Decimal(0.0)
        terminal_cash = Decimal(0.0)

        terminal_totals = {}
        for terminal in terminals_to_template:
            terminal_total_sale += terminal['total_sale']
            terminal_net_sale += terminal['net_sale']
            terminal_vat += terminal['vat']
            terminal_discount += terminal['discount']
            terminal_credit_card += terminal['credit_card']
            terminal_mobile_payment += terminal['mobile_payment']
            terminal_cash += terminal['cash']

        terminal_totals['total_sale'] = terminal_total_sale
        terminal_totals['net_sale'] = terminal_net_sale
        terminal_totals['vat'] = terminal_vat
        terminal_totals['discount'] = terminal_discount
        terminal_totals['credit_card'] = terminal_credit_card
        terminal_totals['mobile_payment'] = terminal_mobile_payment
        terminal_totals['cash'] = terminal_cash

        print(terminal_totals)

        return render(request, 'todays_transaction/todays_transaction.html', {'terminals':terminals_to_template, 'last_update':last_update, 'terminal_totals':terminal_totals})

        

from django.views.generic import ListView
from django.db.models import Q
from .models import Bill, BillItem, Product

class CategoryWiseSale(BillFilterDateMixin,ExportExcelMixin,ListView):
    template_name = "bill/report/category_wise_sales.html"
    queryset = Bill.objects.filter(status=True, is_deleted=False)

    def get_processed_data(self):
        qs = self.get_queryset()
        sales_type = self.request.GET.get('sales_type', 'all')
        if sales_type == "complimentary":
            qs = qs.filter(payment_mode="complimentary")
        elif sales_type == "others":
            qs = qs.filter(~Q(payment_mode="complimentary"))
        
        budclass_data = {}
        
        for bill in qs:
            for item in bill.bill_items.all():
                product_title = item.product.title
                budclass = item.product.category.title
                budclass_items = budclass_data.get(budclass, {'items':[], 'amount_total': 0, 'quantity_total': 0})
                
                item_exists = list(filter(lambda x: (x['name'] == product_title and x['rate'] == item.rate), budclass_items['items']))
                
                if not item_exists:
                    budclass_items['items'].append({'name':product_title,'quantity':item.product_quantity, 'amount':item.amount, 'unit':item.unit_title, 'rate':item.rate})
                    budclass_items['amount_total'] += item.amount
                    budclass_items['quantity_total'] += item.product_quantity
                else:
                    item_exists[0]['quantity'] += item.product_quantity
                    item_exists[0]['amount'] += item.amount
                    budclass_items['amount_total'] += item.amount
                    budclass_items['quantity_total'] += item.product_quantity
        
                budclass_data[budclass] = budclass_items
        
        for k, v in budclass_data.items():
            new_array = sorted(v['items'], key=lambda x: (x['quantity']), reverse=True)
            budclass_data[k]['items'] = new_array
        
        return budclass_data

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        branches = Branch.objects.active()
        context['branches'] = branches
        context['data'] = self.get_processed_data()
        return context

    def export_to_xls(self):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="CategoryWise_report.xls"'

        columns = [
            "type",
            "name",
            "unit",
            "quantity",
            "rate",
            "amount",
        ]
        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Category Wise Sales Report", columns
        )
        
        data = self.get_processed_data()
        new_data = []
        
        for k, v in data.items():
            type_name = k.title()  # Capitalize the first letter of the budclass name
            new_data.append([type_name])
            
            for item in v['items']:
                new_data.append(['', item['name'], item['unit'], item['quantity'], item['rate'], item['amount']])
        
        for row in new_data:
            row_num += 1
            for col_num in range(len(row)):
                value = row[col_num]
                ws.write(row_num, col_num, value, font_style_normal)
        
        wb.save(response)
        return response
        
from bill.models import MobilePaymentType
from bill.forms import MobilePaymentForm
class MobilePaymentTypeMixin(IsAdminMixin):
    model = MobilePaymentType
    form_class = MobilePaymentForm
    paginate_by = 50
    queryset = MobilePaymentType.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("mobile_payment_type_list")
    search_lookup_fields = [
        "name",
        "company",
    ]


class MobilePaymentTypeList(MobilePaymentTypeMixin, ListView):
    template_name = "mobilepaymenttype/mobilepaymenttype_list.html"
    queryset = MobilePaymentType.objects.filter(status=True, is_deleted=False)


class MobilePaymentTypeDetail(MobilePaymentTypeMixin, DetailView):
    template_name = "mobilepaymenttype/mobilepaymenttype_detail.html"


class MobilePaymentTypeCreate(MobilePaymentTypeMixin, CreateView):
    template_name = "create.html"


class MobilePaymentTypeUpdate(MobilePaymentTypeMixin, UpdateView):
    template_name = "update.html"


class MobilePaymentTypeDelete(MobilePaymentTypeMixin, DeleteMixin, View):
    pass


from django.views import View
from django.shortcuts import redirect
from django.contrib import messages
from openpyxl import load_workbook
from django.urls import reverse_lazy
from decimal import Decimal
import nepali_datetime
from .models import TblSalesEntry


from django.views import View
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from decimal import Decimal
from openpyxl import load_workbook
from datetime import datetime
import nepali_datetime

from .models import TblSalesEntry  # Update with correct import path

from django.views import View
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from decimal import Decimal
from openpyxl import load_workbook
from datetime import datetime
import nepali_datetime

from .models import TblSalesEntry


class SalesEntryUploadView(View):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            messages.error(request, "No file uploaded.", extra_tags='danger')
            return redirect(reverse_lazy('tblsalesentry_list'))

        wb = load_workbook(file)
        errors = []

        def clean_decimal(cell_value):
            try:
                return Decimal(str(cell_value).replace(",", "")) if cell_value else Decimal(0)
            except Exception:
                raise ValueError(f"Invalid decimal value: {cell_value}")

        def parse_excel_date(miti_raw, row_idx):
            """Parses B.S. date from Excel either as datetime or string format."""
            if isinstance(miti_raw, datetime):
                d = miti_raw.date()
                if d.year >= 2070:  # Assumed B.S.
                    try:
                        bs_date = nepali_datetime.date(d.year, d.month, d.day)
                        return bs_date.to_datetime_date(), f"{d.year}/{d.month}/{d.day}"
                    except Exception as e:
                        raise ValueError(f"Row {row_idx}: Invalid B.S. datetime object: {d}")
                else:
                    raise ValueError(f"Row {row_idx}: Excel date appears to be in A.D., expected B.S.")
            elif isinstance(miti_raw, str):
                if miti_raw.lower() in ['b.s.', 'miti', 'date']:
                    raise ValueError(f"Row {row_idx}: Skipped header row")
                try:
                    if '-' in miti_raw:
                        parts = list(map(int, miti_raw.split('-')))
                    else:
                        parts = list(map(int, miti_raw.split('/')))
                    bs_date = nepali_datetime.date(parts[0], parts[1], parts[2])
                    return bs_date.to_datetime_date(), miti_raw
                except Exception:
                    raise ValueError(f"Row {row_idx}: Invalid B.S. date string format: {miti_raw}")
            else:
                raise ValueError(f"Row {row_idx}: Unsupported date type: {type(miti_raw)}")

        for sheet in wb.worksheets:
            for idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
                try:
                    miti_raw = row[0].value

                    try:
                        ad_date, miti_bs = parse_excel_date(miti_raw, idx)
                    except Exception as e:
                        errors.append(str(e))
                        continue

                    bill_no = str(row[1].value).strip() if row[1].value else ''
                    customer_name = str(row[2].value).strip() if row[2].value else ''
                    pan_no = str(row[3].value).strip() if row[3].value else ''

                    amount = clean_decimal(row[7].value) + clean_decimal(row[8].value)
                    no_tax_sales = clean_decimal(row[5].value)
                    zero_tax_sales = clean_decimal(row[6].value)
                    taxable_amount = clean_decimal(row[7].value)
                    tax_amount = clean_decimal(row[8].value)

                    TblSalesEntry.objects.create(
                        bill_date=str(ad_date),
                        miti=miti_bs,
                        bill_no=bill_no,
                        customer_name=customer_name,
                        customer_pan=pan_no,
                        amount=amount,
                        NoTaxSales=no_tax_sales,
                        ZeroTaxSales=zero_tax_sales,
                        taxable_amount=taxable_amount,
                        tax_amount=tax_amount
                    )
                    print(f"Row {idx}: Sales entry saved successfully.")

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    print(f"Row {idx} error: {str(e)}")

        if errors:
            messages.error(request, f"Some rows failed to upload:\n {errors}", extra_tags='danger')
        else:
            messages.success(request, "Sales entries uploaded successfully", extra_tags='success')

        return redirect(reverse_lazy('tblsalesentry_list'))
        
from django.views import View
from django.http import JsonResponse

class CreateCustomerFromSalesEntry(View):
    def post(self, request, *args, **kwargs):
        try:
            salesentries = TblSalesEntry.objects.all()
            created_count = 0

            for salesentry in salesentries:
                customer_pan = salesentry.customer_pan
                customer_name = salesentry.customer_name

                if customer_pan:
                    exists = Customer.objects.filter(
                        tax_number=customer_pan,
                        is_deleted=False,
                        status=True
                    ).exists()

                    if not exists:
                        Customer.objects.create(name=customer_name, tax_number=customer_pan)
                        created_count += 1
                else:
                    if not Customer.objects.filter(name=customer_name).exists():
                        Customer.objects.create(name=customer_name)
                        created_count += 1
                    else:
                        continue

            return JsonResponse({"status": "success", "message": f"{created_count} customers created."})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)