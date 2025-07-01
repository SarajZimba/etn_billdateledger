from datetime import date
from django.urls import reverse_lazy
from django.db.models import Sum
from django.views.generic import CreateView,DetailView,ListView,UpdateView,View
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from accounting.models import TblCrJournalEntry, TblDrJournalEntry, TblJournalEntry, AccountLedger, AccountChart, AccountSubLedger, Depreciation
from accounting.utils import calculate_depreciation
from root.utils import DeleteMixin
from product.models import Product
from organization.models import Organization
from product.models import ProductStock, ProductCategory
from .forms import VendorForm, ProductPurchaseForm
from .models import Vendor, ProductPurchase, Purchase, TblpurchaseEntry, TblpurchaseReturn
import decimal
from bill.views import ExportExcelMixin
import json
from django.db.utils import IntegrityError
from user.permission import IsAdminMixin
from bill.utils import create_cumulative_ledger_bill, update_cumulative_ledger_bill

class VendorMixin(IsAdminMixin):
    model = Vendor
    form_class = VendorForm
    paginate_by = 10
    queryset = Vendor.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('vendor_list')



class VendorList(VendorMixin, ListView):
    template_name = "vendor/vendor_list.html"
    queryset = Vendor.objects.filter(status=True,is_deleted=False)


class VendorDetail(VendorMixin, DetailView):
    template_name = "vendor/vendor_detail.html"


class VendorCreate(VendorMixin, CreateView):
    template_name = "create.html"


class VendorUpdate(VendorMixin, UpdateView):
    template_name = "update.html"


class VendorDelete(VendorMixin, DeleteMixin, View):
    pass

'''  -------------------------------------    '''
    
from django.db.models import Q
from accounting.models import AccountSubLedgerTracking
from django.db import transaction

class ProductPurchaseCreateView(IsAdminMixin, CreateView):
    model = ProductPurchase
    form_class = ProductPurchaseForm
    template_name = "purchase/purchase_create.html"

    def create_subledgers(self, product, item_total, debit_account):
        debit_account = get_object_or_404(AccountLedger, pk=int(debit_account))
        subledgername = f'{product.title} ({product.category.title}) - Purchase'
        try:
            sub = AccountSubLedger.objects.get(sub_ledger_name=subledgername, ledger=debit_account)
            prev_value = sub.total_value
            subledgertracking = AccountSubLedgerTracking.objects.create(subledger = sub, prev_amount= prev_value)
            sub.total_value += decimal.Decimal(item_total)
            sub.save()
            subledgertracking.new_amount=sub.total_value
            subledgertracking.value_changed = sub.total_value - prev_value
            subledgertracking.save()
        except AccountSubLedger.DoesNotExist:
            subledger = AccountSubLedger.objects.create(sub_ledger_name=subledgername, ledger=debit_account, total_value=item_total)
            subledgertracking = AccountSubLedgerTracking.objects.create(subledger=subledger, new_amount=decimal.Decimal(item_total), value_changed=decimal.Decimal(item_total))

    def create_accounting_multiple_ledger(self, debit_account_id, payment_mode:str, username:str, sub_total, tax_amount, vendor, excise_duty_amount):
        sub_total = decimal.Decimal(sub_total)
        tax_amount = decimal.Decimal(tax_amount)
        total_amount =  sub_total+ tax_amount + excise_duty_amount

        cash_ledger = get_object_or_404(AccountLedger, ledger_name='Cash-In-Hand')
        vat_receivable = get_object_or_404(AccountLedger, ledger_name='VAT Receivable')
        excise_duty_receivable = get_object_or_404(AccountLedger, ledger_name='Excise Duty Receivable')
        debit_account = get_object_or_404(AccountLedger, pk=int(debit_account_id))
        
        journal_entry = TblJournalEntry.objects.create(employee_name=username, journal_total = total_amount)
        TblDrJournalEntry.objects.create(journal_entry=journal_entry, particulars=f"Automatic: {debit_account.ledger_name} A/c Dr.", debit_amount=sub_total, ledger=debit_account)
        debit_account.total_value += sub_total
        debit_account.save()
        update_cumulative_ledger_bill(debit_account)
        if tax_amount > 0:
            TblDrJournalEntry.objects.create(journal_entry=journal_entry, particulars="Automatic: VAT Receivable A/c Dr.", debit_amount=tax_amount, ledger=vat_receivable)
            vat_receivable.total_value += tax_amount
            vat_receivable.save()
            update_cumulative_ledger_bill(vat_receivable)
        if excise_duty_amount > 0:
            TblDrJournalEntry.objects.create(journal_entry=journal_entry, particulars="Automatic: Excise Duty Receivable A/c Dr.", debit_amount=excise_duty_amount, ledger=excise_duty_receivable)
            excise_duty_receivable.total_value += excise_duty_amount
            excise_duty_receivable.save()
            update_cumulative_ledger_bill(excise_duty_receivable)
        if payment_mode.lower().strip() == "credit":
            try:
                vendor_ledger = AccountLedger.objects.get(ledger_name=vendor)
                vendor_ledger.total_value += total_amount
                vendor_ledger.save()
                update_cumulative_ledger_bill(vendor_ledger)
            except AccountLedger.DoesNotExist:
                chart = AccountChart.objects.get(group__iexact='Sundry Creditors')
                vendor_ledger = AccountLedger.objects.create(ledger_name=vendor, total_value=total_amount, is_editable=True, account_chart=chart)
                create_cumulative_ledger_bill(vendor_ledger)
            TblCrJournalEntry.objects.create(journal_entry=journal_entry, particulars=f"Automatic: To {vendor_ledger.ledger_name} A/c", credit_amount=total_amount, ledger=vendor_ledger)
        elif payment_mode.lower().strip() == "mobile payment":
            # Get required ledgers
            mobile_payment_ledger = get_object_or_404(AccountLedger, ledger_name='Mobile Payments')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {mobile_payment_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=mobile_payment_ledger
            )
            mobile_payment_ledger.total_value -= total_amount
            mobile_payment_ledger.save()
            update_cumulative_ledger_bill(mobile_payment_ledger)
        elif payment_mode.lower().strip() == "credit card":
            # Get required ledgers
            credit_card_ledger = get_object_or_404(AccountLedger, ledger_name='Card Transactions')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {credit_card_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=credit_card_ledger
            )
            credit_card_ledger.total_value -= total_amount
            credit_card_ledger.save()
            update_cumulative_ledger_bill(credit_card_ledger)
        elif payment_mode.lower().strip() == "complimentary":
            # Get required ledgers
            complimentary_expense_ledger = get_object_or_404(AccountLedger, ledger_name='Complimentary Expenses')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {complimentary_expense_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=complimentary_expense_ledger
            )
            complimentary_expense_ledger.total_value -= total_amount
            complimentary_expense_ledger.save()
            update_cumulative_ledger_bill(complimentary_expense_ledger)
        else:
            TblCrJournalEntry.objects.create(journal_entry=journal_entry, particulars=f"Automatic: To {cash_ledger.ledger_name} A/c", credit_amount=total_amount, ledger=cash_ledger)
            cash_ledger.total_value -= total_amount
            cash_ledger.save()
            update_cumulative_ledger_bill(cash_ledger)

    def create_accounting_single_ledger(self, ledger_totals, payment_mode: str, username: str, 
                                    tax_amount: decimal.Decimal, excise_duty_amount: decimal.Decimal, 
                                    vendor: str):
        """
        Create a single journal entry for all product ledgers
        ledger_totals: dict of {ledger_id: total_amount}
        """
        # Calculate grand total
        sub_total = sum(decimal.Decimal(total) for total in ledger_totals.values())
        total_amount = sub_total + tax_amount + excise_duty_amount

        # Get required ledgers
        cash_ledger = get_object_or_404(AccountLedger, ledger_name='Cash-In-Hand')
        vat_receivable = get_object_or_404(AccountLedger, ledger_name='VAT Receivable')
        excise_duty_receivable = get_object_or_404(AccountLedger, ledger_name='Excise Duty Receivable')
        
        # Create single journal entry
        journal_entry = TblJournalEntry.objects.create(
            employee_name=username, 
            journal_total=total_amount
        )
        
        # Create debit entries for each product ledger
        for ledger_id, amount in ledger_totals.items():
            ledger = get_object_or_404(AccountLedger, pk=int(ledger_id))
            TblDrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: {ledger.ledger_name} A/c Dr.",
                debit_amount=decimal.Decimal(amount),
                ledger=ledger
            )
            # Update ledger balance
            ledger.total_value += decimal.Decimal(amount)
            ledger.save()
            update_cumulative_ledger_bill(ledger)
        
        # Add tax and excise duty entries if applicable
        if tax_amount > 0:
            TblDrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars="Automatic: VAT Receivable A/c Dr.",
                debit_amount=tax_amount,
                ledger=vat_receivable
            )
            vat_receivable.total_value += tax_amount
            vat_receivable.save()
            update_cumulative_ledger_bill(vat_receivable)
        
        if excise_duty_amount > 0:
            TblDrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars="Automatic: Excise Duty Receivable A/c Dr.",
                debit_amount=excise_duty_amount,
                ledger=excise_duty_receivable
            )
            excise_duty_receivable.total_value += excise_duty_amount
            excise_duty_receivable.save()
            update_cumulative_ledger_bill(excise_duty_receivable)
        
        # Create credit entry
        if payment_mode.lower().strip() == "credit":
            try:
                vendor_ledger = AccountLedger.objects.get(ledger_name=vendor)
                vendor_ledger.total_value += total_amount
                vendor_ledger.save()
                update_cumulative_ledger_bill(vendor_ledger)
            except AccountLedger.DoesNotExist:
                chart = AccountChart.objects.get(group__iexact='Sundry Creditors')
                vendor_ledger = AccountLedger.objects.create(
                    ledger_name=vendor,
                    total_value=total_amount,
                    is_editable=True,
                    account_chart=chart
                )
                create_cumulative_ledger_bill(vendor_ledger)
            

            
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {vendor_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=vendor_ledger
            )
        elif payment_mode.lower().strip() == "mobile payment":
            # Get required ledgers
            mobile_payment_ledger = get_object_or_404(AccountLedger, ledger_name='Mobile Payments')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {mobile_payment_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=mobile_payment_ledger
            )
            mobile_payment_ledger.total_value -= total_amount
            mobile_payment_ledger.save()
            update_cumulative_ledger_bill(mobile_payment_ledger)
        elif payment_mode.lower().strip() == "credit card":
            # Get required ledgers
            credit_card_ledger = get_object_or_404(AccountLedger, ledger_name='Card Transactions')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {credit_card_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=credit_card_ledger
            )
            credit_card_ledger.total_value -= total_amount
            credit_card_ledger.save()
            update_cumulative_ledger_bill(credit_card_ledger)
        elif payment_mode.lower().strip() == "complimentary":
            # Get required ledgers
            complimentary_expense_ledger = get_object_or_404(AccountLedger, ledger_name='Complimentary Expenses')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {complimentary_expense_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=complimentary_expense_ledger
            )
            complimentary_expense_ledger.total_value -= total_amount
            complimentary_expense_ledger.save()
            update_cumulative_ledger_bill(complimentary_expense_ledger)
        else:
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {cash_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=cash_ledger
            )
            cash_ledger.total_value -= total_amount
            cash_ledger.save()
            update_cumulative_ledger_bill(cash_ledger)

    def form_invalid(self, form) -> HttpResponse:
        return self.form_valid(form)
        
    @transaction.atomic()    
    def form_valid(self, form):
        form_data = form.data 
        print(form.data)
        bill_no = form_data.get('bill_no', None)
        bill_date = form_data.get('bill_date', None)
        pp_no = form_data.get('pp_no',None)
        vendor_id = form_data.get('vendor')
        sub_total = form_data.get('sub_total')
        discount_percentage = form_data.get('discount_percentage')
        discount_amount = form_data.get('discount_amount')
        taxable_amount = form_data.get('taxable_amount')
        non_taxable_amount = form_data.get('non_taxable_amount')
        tax_amount = form_data.get('tax_amount')
        grand_total = form_data.get('grand_total')
        amount_in_words = form_data.get('amount_in_words')
        payment_mode = form_data.get('payment_mode')
        debit_account = form_data.get('debit_account')
        excise_duty_amount = form_data.get('excise_duty_amount')
        # print(debit_account)
        purchase_object = Purchase(
            bill_no=bill_no,
            vendor_id=vendor_id,sub_total=sub_total, bill_date=bill_date,
            discount_percentage=discount_percentage,discount_amount=discount_amount,
            taxable_amount=taxable_amount, non_taxable_amount=non_taxable_amount,
            tax_amount=tax_amount, grand_total=grand_total,
            amount_in_words=amount_in_words, payment_mode=payment_mode,
            excise_duty_amount=decimal.Decimal(excise_duty_amount)
        )
        purchase_object.save()

        product_ids =  form_data.get('product_id_list', '')
        product_taxable_info = form_data.get('product_taxable_info', '')
        product_ledger_info = form_data.get('ledger_id_list', '')
        product_ledger_info_parse = json.loads(product_ledger_info)
        # print(product_ledger_info)
        no_of_items_sent = len(product_ledger_info_parse)
        product_category_info = form_data.get('product_category_info')
        print(product_category_info)


        new_items_name = {}
        new_product_categories = {}
        new_product_ledgers = {}
        if product_taxable_info and len(product_taxable_info) > 0:
            new_items_name = json.loads(product_taxable_info)
            # print(new_items_name)


            new_product_categories = json.loads(product_category_info)
            new_product_ledgers = json.loads(product_ledger_info)

        item_name = ''

        total_quantity = 0
        vendor = Vendor.objects.get(pk=vendor_id)
        vendor_name = vendor.name
        vendor_pan = vendor.pan_no

        if product_ids:
            product_ids = product_ids.split(',')


        
        if product_ledger_info and len(product_ledger_info) > 0:
            product_ledgers = json.loads(product_ledger_info)
            
            for product_id, ledger_info in product_ledgers.items():
                try:
                    product_id = int(product_id)
                    ledger_id = int(ledger_info['ledgerId'])
                    total = float(ledger_info['total'])
                    # print(product_id)
                    # print(ledger_id)
                    
                    quantity = float(form_data.get(f'id_bill_item_quantity_{product_id}'))
                    rate = float(form_data.get(f'id_bill_item_rate_{product_id}'))
                    item_total = quantity * rate
                    # print(quantity)
                    # print(rate)

                    # Get the product and ledger objects
                    prod = Product.objects.get(pk=product_id)
                    ledger = AccountLedger.objects.get(pk=ledger_id)

                    # Debit the ledger for the product
                    self.create_subledgers(prod, item_total, ledger_id)
                    ProductPurchase.objects.create(product=prod, purchase=purchase_object, quantity=quantity, rate=rate, item_total=total)
                    from organization.models import Branch
                    from product.models import BranchStock
                    BranchStock.objects.create(branch=Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first(), product=prod, quantity=quantity)
                except (ValueError, Product.DoesNotExist, AccountLedger.DoesNotExist):
                    pass
        


        if new_items_name:
            for k, v in new_items_name.items():
                category_name = new_product_categories.get(k, '').lower().strip()
                # print(category_name)
                # ledger_name = new_product_ledgers
                if ProductCategory.objects.filter(title__iexact=category_name).exists():
                    category = ProductCategory.objects.filter(title__iexact=category_name).first()
                else:
                    try:
                        ProductCategory.objects.create(title=category_name)
                    except IntegrityError:
                        pass
                category = ProductCategory.objects.filter(title__iexact=category_name).first()
                rate = float(form_data.get(f'id_bill_item_rate_{k}'))
                quantity = float(form_data.get(f'id_bill_item_quantity_{k}'))
                item_total = quantity * rate
                is_taxable = True if (v == "true" or v == True) else False
                ledger_info = json.loads(product_ledger_info)
                # print(ledger_info)
                ledger_id = int(ledger_info.get(k, {}).get('ledgerId', ''))
                ledger = AccountLedger.objects.get(id=ledger_id)
                # print(ledger)
                clean_title = k.replace('-', ' ')
                try:
                    prod = Product.objects.create(category=category, title=clean_title, is_taxable=is_taxable, price=rate, ledger=ledger, is_billing_item = False)
                except IntegrityError:
                    prod = Product.objects.get(title__iexact=k)
                self.create_subledgers(prod, item_total, ledger_id)
                ProductPurchase.objects.create(product=prod, purchase=purchase_object, quantity=quantity, rate=rate, item_total=item_total)
                from organization.models import Branch
                from product.models import BranchStock
                BranchStock.objects.create(branch=Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first(), product=prod, quantity=quantity)
        TblpurchaseEntry.objects.create(
            bill_no=bill_no, bill_date=bill_date, pp_no=pp_no, vendor_name=vendor_name, vendor_pan=vendor_pan,
            item_name=item_name, quantity=total_quantity, amount=grand_total, tax_amount=tax_amount, non_tax_purchase=non_taxable_amount, excise_duty_amount=decimal.Decimal(excise_duty_amount), purchase_id=purchase_object.id, taxable_amount=taxable_amount
        )
        vendor_detail = str(vendor.pk)+' '+ vendor_name
        # self.create_accounting(debit_account_id=debit_account, payment_mode=payment_mode, username=self.request.user.username, sub_total=sub_total, tax_amount=tax_amount, vendor=vendor_detail)
        sub_tax = decimal.Decimal(tax_amount)
        fraction_tax = sub_tax/no_of_items_sent
        sub_excise_duty = decimal.Decimal(excise_duty_amount)
        fraction_excise_duty = sub_excise_duty/no_of_items_sent
        print(fraction_tax)
        # if product_ledger_info and len(product_ledger_info) > 0:
        #     product_ledgers = json.loads(product_ledger_info)
            
        #     for product_id, ledger_info in product_ledgers.items():
        #         ledger_id = int(ledger_info['ledgerId'])
        #         total = float(ledger_info['total'])
        #         self.create_accounting_multiple_ledger(debit_account_id=ledger_id, payment_mode=payment_mode, username=self.request.user.username, sub_total=total, tax_amount=fraction_tax, vendor=vendor_detail, excise_duty_amount=fraction_excise_duty)
        # Collect all ledger totals first
        ledger_totals = {}
        if product_ledger_info and len(product_ledger_info) > 0:
            product_ledgers = json.loads(product_ledger_info)
            for product_id, ledger_info in product_ledgers.items():
                ledger_id = ledger_info['ledgerId']
                total = float(ledger_info['total'])
                if ledger_id in ledger_totals:
                    ledger_totals[ledger_id] += total
                else:
                    ledger_totals[ledger_id] = total
        
        # Create single journal entry with all ledgers
        if ledger_totals:
            self.create_accounting_single_ledger(
                ledger_totals=ledger_totals,
                payment_mode=payment_mode,
                username=self.request.user.username,
                tax_amount=decimal.Decimal(tax_amount),
                excise_duty_amount=decimal.Decimal(excise_duty_amount),
                vendor=vendor_detail
            )

        return redirect('/purchase/')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Fetch ledgers with the account_chart of purchases and expenses
        purchases_and_expenses_ledgers = AccountLedger.objects.filter(
            Q(account_chart__account_type='Expense') | Q(account_chart__group='Purchases')
        )
        # print(purchases_and_expenses_ledgers)

        # Add the fetched ledgers to the context
        context['purchases_and_expenses_ledgers'] = purchases_and_expenses_ledgers

        return context

# class ProductPurchaseCreateView(IsAdminMixin, CreateView):
#     model = ProductPurchase
#     form_class = ProductPurchaseForm
#     template_name = "purchase/purchase_create.html"

#     def create_subledgers(self, product, item_total, debit_account):
#         debit_account = get_object_or_404(AccountLedger, pk=int(debit_account))
#         subledgername = f'{product.title} ({product.category.title}) - Purchase'
#         try:
#             sub = AccountSubLedger.objects.get(sub_ledger_name=subledgername, ledger=debit_account)
#             prev_value = sub.total_value
#             subledgertracking = AccountSubLedgerTracking.objects.create(subledger = sub, prev_amount= prev_value)
#             sub.total_value += decimal.Decimal(item_total)
#             sub.save()
#             subledgertracking.new_amount=sub.total_value
#             subledgertracking.value_changed = sub.total_value - prev_value
#             subledgertracking.save()
#         except AccountSubLedger.DoesNotExist:
#             subledger = AccountSubLedger.objects.create(sub_ledger_name=subledgername, ledger=debit_account, total_value=item_total)
#             subledgertracking = AccountSubLedgerTracking.objects.create(subledger=subledger, new_amount=decimal.Decimal(item_total), value_changed=decimal.Decimal(item_total))

#     def create_accounting_single_ledger(self, ledger_totals, payment_mode: str, username: str, 
#                                     tax_amount: decimal.Decimal, 
#                                     vendor: str, entry_date):
#         """
#         Create a single journal entry for all product ledgers
#         ledger_totals: dict of {ledger_id: total_amount}
#         """
#         # Calculate grand total
#         sub_total = sum(decimal.Decimal(total) for total in ledger_totals.values())
#         total_amount = sub_total + tax_amount 

#         # Get required ledgers
#         cash_ledger = get_object_or_404(AccountLedger, ledger_name='Cash-In-Hand')
#         vat_receivable = get_object_or_404(AccountLedger, ledger_name='VAT Receivable')
        
#         # Create single journal entry
#         journal_entry = TblJournalEntry.objects.create(
#             employee_name=username, 
#             journal_total=total_amount,
#             entry_date=entry_date
#         )

#         if entry_date:

#             entry_datetime_for_cumulativeledger = change_date_to_datetime(entry_date)
#         else:
#             from datetime import datetime
#             entry_datetime_for_cumulativeledger = datetime.now()

#         # Create debit entries for each product ledger
#         for ledger_id, amount in ledger_totals.items():
#             ledger = get_object_or_404(AccountLedger, pk=int(ledger_id))
#             TblDrJournalEntry.objects.create(
#                 journal_entry=journal_entry,
#                 particulars=f"Automatic: {ledger.ledger_name} A/c Dr.",
#                 debit_amount=decimal.Decimal(amount),
#                 ledger=ledger
#             )
#             # Update ledger balance
#             ledger.total_value += decimal.Decimal(amount)
#             ledger.save()
#             update_cumulative_ledger_bill(ledger, entry_datetime_for_cumulativeledger)
        
#         # Add tax and excise duty entries if applicable
#         if tax_amount > 0:
#             TblDrJournalEntry.objects.create(
#                 journal_entry=journal_entry,
#                 particulars="Automatic: VAT Receivable A/c Dr.",
#                 debit_amount=tax_amount,
#                 ledger=vat_receivable
#             )
#             vat_receivable.total_value += tax_amount
#             vat_receivable.save()
#             update_cumulative_ledger_bill(vat_receivable, entry_datetime_for_cumulativeledger)
        
#         # Create credit entry
#         if payment_mode.lower().strip() == "credit":
#             try:
#                 vendor_ledger = AccountLedger.objects.get(ledger_name=vendor)
#                 vendor_ledger.total_value += total_amount
#                 vendor_ledger.save()
#                 update_cumulative_ledger_bill(vendor_ledger, entry_datetime_for_cumulativeledger)
#             except AccountLedger.DoesNotExist:
#                 chart = AccountChart.objects.get(group__iexact='Sundry Creditors')
#                 vendor_ledger = AccountLedger.objects.create(
#                     ledger_name=vendor,
#                     total_value=total_amount,
#                     is_editable=True,
#                     account_chart=chart
#                 )
#                 create_cumulative_ledger_bill(vendor_ledger, entry_datetime_for_cumulativeledger)
            

            
#             TblCrJournalEntry.objects.create(
#                 journal_entry=journal_entry,
#                 particulars=f"Automatic: To {vendor_ledger.ledger_name} A/c",
#                 credit_amount=total_amount,
#                 ledger=vendor_ledger
#             )
#         elif payment_mode.lower().strip() == "mobile payment":
#             # Get required ledgers
#             mobile_payment_ledger = get_object_or_404(AccountLedger, ledger_name='Mobile Payments')
#             TblCrJournalEntry.objects.create(
#                 journal_entry=journal_entry,
#                 particulars=f"Automatic: To {mobile_payment_ledger.ledger_name} A/c",
#                 credit_amount=total_amount,
#                 ledger=mobile_payment_ledger
#             )
#             mobile_payment_ledger.total_value -= total_amount
#             mobile_payment_ledger.save()
#             update_cumulative_ledger_bill(mobile_payment_ledger, entry_datetime_for_cumulativeledger)
#         elif payment_mode.lower().strip() == "credit card":
#             # Get required ledgers
#             credit_card_ledger = get_object_or_404(AccountLedger, ledger_name='Card Transactions')
#             TblCrJournalEntry.objects.create(
#                 journal_entry=journal_entry,
#                 particulars=f"Automatic: To {credit_card_ledger.ledger_name} A/c",
#                 credit_amount=total_amount,
#                 ledger=credit_card_ledger
#             )
#             credit_card_ledger.total_value -= total_amount
#             credit_card_ledger.save()
#             update_cumulative_ledger_bill(credit_card_ledger, entry_datetime_for_cumulativeledger)
#         elif payment_mode.lower().strip() == "complimentary":
#             # Get required ledgers
#             complimentary_expense_ledger = get_object_or_404(AccountLedger, ledger_name='Complimentary Expenses')
#             TblCrJournalEntry.objects.create(
#                 journal_entry=journal_entry,
#                 particulars=f"Automatic: To {complimentary_expense_ledger.ledger_name} A/c",
#                 credit_amount=total_amount,
#                 ledger=complimentary_expense_ledger
#             )
#             complimentary_expense_ledger.total_value -= total_amount
#             complimentary_expense_ledger.save()
#             update_cumulative_ledger_bill(complimentary_expense_ledger, entry_datetime_for_cumulativeledger)
#         else:
#             TblCrJournalEntry.objects.create(
#                 journal_entry=journal_entry,
#                 particulars=f"Automatic: To {cash_ledger.ledger_name} A/c",
#                 credit_amount=total_amount,
#                 ledger=cash_ledger
#             )
#             cash_ledger.total_value -= total_amount
#             cash_ledger.save()
#             update_cumulative_ledger_bill(cash_ledger, entry_datetime_for_cumulativeledger)

#     def form_invalid(self, form) -> HttpResponse:
#         return self.form_valid(form)
#     @transaction.atomic()
#     def form_valid(self, form):
#         form_data = form.data 
#         print(form.data)
#         bill_no = form_data.get('bill_no', None)
#         bill_date = form_data.get('bill_date', None)
#         pp_no = form_data.get('pp_no',None)
#         vendor_id = form_data.get('vendor')
#         sub_total = form_data.get('sub_total')
#         discount_percentage = form_data.get('discount_percentage')
#         discount_amount = form_data.get('discount_amount')
#         taxable_amount = form_data.get('taxable_amount')
#         non_taxable_amount = form_data.get('non_taxable_amount')
#         tax_amount = form_data.get('tax_amount')
#         grand_total = form_data.get('grand_total')
#         amount_in_words = form_data.get('amount_in_words')
#         payment_mode = form_data.get('payment_mode')
#         debit_account = form_data.get('debit_account')
#         # print(debit_account)
#         purchase_object = Purchase(
#             bill_no=bill_no,
#             vendor_id=vendor_id,sub_total=sub_total, bill_date=bill_date,
#             discount_percentage=discount_percentage,discount_amount=discount_amount,
#             taxable_amount=taxable_amount, non_taxable_amount=non_taxable_amount,
#             tax_amount=tax_amount, grand_total=grand_total,
#             amount_in_words=amount_in_words, payment_mode=payment_mode
#         )
#         purchase_object.save()

#         product_ids =  form_data.get('product_id_list', '')
#         product_taxable_info = form_data.get('product_taxable_info', '')
#         product_ledger_info = form_data.get('ledger_id_list', '')
#         product_ledger_info_parse = json.loads(product_ledger_info)
#         # print(product_ledger_info)
#         no_of_items_sent = len(product_ledger_info_parse)
#         product_category_info = form_data.get('product_category_info')
#         print(product_category_info)


#         new_items_name = {}
#         new_product_categories = {}
#         new_product_ledgers = {}
#         if product_taxable_info and len(product_taxable_info) > 0:
#             new_items_name = json.loads(product_taxable_info)
#             # print(new_items_name)


#             new_product_categories = json.loads(product_category_info)
#             new_product_ledgers = json.loads(product_ledger_info)

#         item_name = ''

#         total_quantity = 0
#         vendor = Vendor.objects.get(pk=vendor_id)
#         vendor_name = vendor.name
#         vendor_pan = vendor.pan_no

#         if product_ids:
#             product_ids = product_ids.split(',')


        
#         if product_ledger_info and len(product_ledger_info) > 0:
#             product_ledgers = json.loads(product_ledger_info)
            
#             for product_id, ledger_info in product_ledgers.items():
#                 try:
#                     product_id = int(product_id)
#                     ledger_id = int(ledger_info['ledgerId'])
#                     total = float(ledger_info['total'])
#                     # print(product_id)
#                     # print(ledger_id)
                    
#                     quantity = float(form_data.get(f'id_bill_item_quantity_{product_id}'))
#                     rate = float(form_data.get(f'id_bill_item_rate_{product_id}'))
#                     item_total = quantity * rate
#                     # print(quantity)
#                     # print(rate)

#                     # Get the product and ledger objects
#                     prod = Product.objects.get(pk=product_id)
#                     ledger = AccountLedger.objects.get(pk=ledger_id)

#                     # Debit the ledger for the product
#                     self.create_subledgers(prod, item_total, ledger_id)
#                     ProductPurchase.objects.create(product=prod, purchase=purchase_object, quantity=quantity, rate=rate, item_total=total)
#                     # from organization.models import Branch
#                     # from product.models import BranchStock
#                     # BranchStock.objects.create(branch=Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first(), product=prod, quantity=quantity)
#                 except (ValueError, Product.DoesNotExist, AccountLedger.DoesNotExist):
#                     pass
        


#         if new_items_name:
#             for k, v in new_items_name.items():
#                 category_name = new_product_categories.get(k, '').lower().strip()
#                 # print(category_name)
#                 # ledger_name = new_product_ledgers
#                 if ProductCategory.objects.filter(title__iexact=category_name).exists():
#                     category = ProductCategory.objects.filter(title__iexact=category_name).first()
#                 else:
#                     try:
#                         ProductCategory.objects.create(title=category_name)
#                     except IntegrityError:
#                         pass
#                 category = ProductCategory.objects.filter(title__iexact=category_name).first()
#                 rate = float(form_data.get(f'id_bill_item_rate_{k}'))
#                 quantity = float(form_data.get(f'id_bill_item_quantity_{k}'))
#                 item_total = quantity * rate
#                 is_taxable = True if (v == "true" or v == True) else False
#                 ledger_info = json.loads(product_ledger_info)
#                 # print(ledger_info)
#                 ledger_id = int(ledger_info.get(k, {}).get('ledgerId', ''))
#                 ledger = AccountLedger.objects.get(id=ledger_id)
#                 # print(ledger)
#                 clean_title = k.replace('-', ' ')
#                 try:
#                     prod = Product.objects.create(category=category, title=clean_title, is_taxable=is_taxable, price=rate, ledger=ledger, is_billing_item = False)
#                 except IntegrityError:
#                     prod = Product.objects.get(title__iexact=k)
#                 self.create_subledgers(prod, item_total, ledger_id)
#                 ProductPurchase.objects.create(product=prod, purchase=purchase_object, quantity=quantity, rate=rate, item_total=item_total)
#                 # from organization.models import Branch
#                 # from product.models import BranchStock
#                 # BranchStock.objects.create(branch=Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first(), product=prod, quantity=quantity)
#         TblpurchaseEntry.objects.create(
#             bill_no=bill_no, bill_date=bill_date, pp_no=pp_no, vendor_name=vendor_name, vendor_pan=vendor_pan,
#             item_name=item_name, quantity=total_quantity, amount=grand_total, tax_amount=tax_amount, non_tax_purchase=non_taxable_amount, purchase_id=purchase_object.id
#         )
#         vendor_detail = str(vendor.pk)+' '+ vendor_name
#         # self.create_accounting(debit_account_id=debit_account, payment_mode=payment_mode, username=self.request.user.username, sub_total=sub_total, tax_amount=tax_amount, vendor=vendor_detail)
#         sub_tax = decimal.Decimal(tax_amount)
#         fraction_tax = sub_tax/no_of_items_sent
#         print(fraction_tax)
#         # if product_ledger_info and len(product_ledger_info) > 0:
#         #     product_ledgers = json.loads(product_ledger_info)
            
#         #     for product_id, ledger_info in product_ledgers.items():
#         #         ledger_id = int(ledger_info['ledgerId'])
#         #         total = float(ledger_info['total'])
#         #         self.create_accounting_multiple_ledger(debit_account_id=ledger_id, payment_mode=payment_mode, username=self.request.user.username, sub_total=total, tax_amount=fraction_tax, vendor=vendor_detail, excise_duty_amount=fraction_excise_duty)
#         # Collect all ledger totals first
#         ledger_totals = {}
#         if product_ledger_info and len(product_ledger_info) > 0:
#             product_ledgers = json.loads(product_ledger_info)
#             for product_id, ledger_info in product_ledgers.items():
#                 ledger_id = ledger_info['ledgerId']
#                 total = float(ledger_info['total'])
#                 if ledger_id in ledger_totals:
#                     ledger_totals[ledger_id] += total
#                 else:
#                     ledger_totals[ledger_id] = total
        
#         # Create single journal entry with all ledgers
#         if ledger_totals:
#             self.create_accounting_single_ledger(
#                 ledger_totals=ledger_totals,
#                 payment_mode=payment_mode,
#                 username=self.request.user.username,
#                 tax_amount=decimal.Decimal(tax_amount),
#                 vendor=vendor_detail,
#                 entry_date=bill_date
#             )

#         return redirect('/purchase/')
    
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)

#         # Fetch ledgers with the account_chart of purchases and expenses
#         purchases_and_expenses_ledgers = AccountLedger.objects.filter(
#             Q(account_chart__account_type='Expense') | Q(account_chart__group='Purchases')
#         )
#         # print(purchases_and_expenses_ledgers)

#         # Add the fetched ledgers to the context
#         context['purchases_and_expenses_ledgers'] = purchases_and_expenses_ledgers

#         return context

class PurchaseListView(IsAdminMixin, ListView):
    model = Purchase
    queryset = Purchase.objects.filter(is_deleted=False)
    template_name = 'purchase/purchase_list.html'


class PurchaseDetailView(IsAdminMixin, DetailView):
    template_name = 'purchase/purchase_detail.html'
    queryset = Purchase.objects.filter(is_deleted=False)

    def get_context_data(self, **kwargs):
        org = Organization.objects.first()
        context =  super().get_context_data(**kwargs)
        context['organization'] = org
        return context


from django.db import transaction
# class MarkPurchaseVoid(IsAdminMixin, View):
#     @transaction.atomic()
#     def post(self, request, *args, **kwargs):
#         id = self.kwargs.get('pk')
#         reason = request.POST.get('voidReason')
#         purchase = get_object_or_404(Purchase, pk=id)
#         purchase.status = False
#         purchase.save()


#         purchased_products = purchase.productpurchase_set.all()
#         for item in purchased_products:
#             stock = ProductStock.objects.get(product=item.product)
#             stock.stock_quantity = stock.stock_quantity-item.quantity
#             stock.save()
            

#         entry_obj = TblpurchaseEntry.objects.get(purchase_id=id)
#         TblpurchaseReturn.objects.create(
#             bill_date=entry_obj.bill_date,
#             bill_no=entry_obj.bill_no,
#             pp_no=entry_obj.pp_no,
#             vendor_name=entry_obj.vendor_name,
#             vendor_pan=entry_obj.vendor_pan,
#             item_name=entry_obj.item_name,
#             quantity=entry_obj.quantity,
#             unit=entry_obj.unit,
#             amount=entry_obj.amount,
#             tax_amount=entry_obj.tax_amount,
#             non_tax_purchase=entry_obj.non_tax_purchase,
#             reason = reason,
#             excise_duty_amount = entry_obj.excise_duty_amount

#         )
        
        
#         return redirect(
#             reverse_lazy("purchase_detail", kwargs={"pk": id})
#         )

class MarkPurchaseVoid(IsAdminMixin, View):
    @transaction.atomic()
    def post(self, request, *args, **kwargs):
        id = self.kwargs.get('pk')
        reason = request.POST.get('voidReason')
        purchase = get_object_or_404(Purchase, pk=id)
        purchase.status = False
        purchase.save()

        purchased_products = purchase.productpurchase_set.all()

        org = Organization.objects.first()
        allow_negative_sales = org.allow_negative_sales
        
        from organization.models import Branch
        from product.models import BranchStock
        
        central_branch = Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first()
        print(purchased_products)
        for product_purchase in purchased_products:
            quantity = product_purchase.quantity
            product = product_purchase.product
            
            # Get all stock entries for this product in central branch, ordered by quantity (ascending)
            branchstock_entries = BranchStock.objects.filter(
                    branch=central_branch, 
                    product=product, 
                    is_deleted=False
                ).order_by('quantity')
                
            remaining_quantity = quantity
                
            for branchstock in branchstock_entries:
                if remaining_quantity <= 0:
                    break
                        
                available_quantity = branchstock.quantity
                    
                if remaining_quantity >= available_quantity:
                    # Delete the entry if quantity becomes zero
                    branchstock.quantity = 0
                    branchstock.is_deleted = True
                    branchstock.save()
                    remaining_quantity -= available_quantity
                else:
                    # Reduce the quantity
                    branchstock.quantity -= remaining_quantity
                    branchstock.save()
                    remaining_quantity = 0

        entry_obj = TblpurchaseEntry.objects.get(purchase_id=id)
        TblpurchaseReturn.objects.create(
            bill_date=entry_obj.bill_date,
            bill_no=entry_obj.bill_no,
            pp_no=entry_obj.pp_no,
            vendor_name=entry_obj.vendor_name,
            vendor_pan=entry_obj.vendor_pan,
            item_name=entry_obj.item_name,
            quantity=entry_obj.quantity,
            unit=entry_obj.unit,
            amount=entry_obj.amount,
            tax_amount=entry_obj.tax_amount,
            taxable_amount=entry_obj.taxable_amount,
            non_tax_purchase=entry_obj.non_tax_purchase,
            reason=reason,
            excise_duty_amount=entry_obj.excise_duty_amount
        )
        
        return redirect(reverse_lazy("purchase_detail", kwargs={"pk": id}))


""" View starting for Purchase Book  """

class PurchaseBookListView(IsAdminMixin, ExportExcelMixin,View):

    # def export_to_excel(self, data):
    #     response = HttpResponse(content_type="application/ms-excel")
    #     response["Content-Disposition"] = 'attachment; filename="purchase_book.xls"'

    #     common = ['bill_date', "bill_no", "pp_no", "vendor_name", "vendor_pan", "amount", "tax_amount", "non_tax_purchase"]
    #     common.insert(0, 'idtblpurchaseEntry')
    #     extra = ["import","importCountry","importNumber", "importDate"]
        

    #     wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
    #         "Purchase Book", common+extra
    #     )
    #     purchase_entry = data.get('purchase_entry')
    #     rows = purchase_entry.values_list(*common)

    #     for row in rows:
    #         row = row + (0,0,0,0)
    #         row_num += 1
    #         for col_num in range(len(row)):
    #             ws.write(row_num, col_num, row[col_num], font_style_normal)

    #     purchase_entry_sum = data.get('purchase_entry_sum')
    #     print(purchase_entry_sum)

    #     row_num += 1
    #     ws.write(row_num, 0, "Total", font_style_normal)
    #     for key, value in purchase_entry_sum.items():
    #         key = key.split('__')[0]
    #         ws.write(row_num, common.index(key), value or 0, font_style_normal)

    #     common [0] = "idtblpurchaseReturn"
    #     columns2 = common+extra

    #     row_num += 1
    #     ws.write(row_num, 0, "")
    #     row_num += 1
    #     ws.write(row_num, 0, "Purchase Return", font_style_bold)
    #     row_num += 1

    #     new_columns = ["id"] + columns2[1:]
    #     for col_num in range(len(columns2)):
    #         ws.write(row_num, col_num, new_columns[col_num], font_style_bold)

    #     return_entry = data.get('return_entry')
    #     rows2 = return_entry.values_list(*common)
    #     return_entry_sum = data.get('return_entry_sum')

    #     for row in rows2:
    #         row = row + (0,0,0,0)
    #         row_num += 1
    #         for col_num in range(len(row)):
    #             ws.write(row_num, col_num, row[col_num], font_style_normal)

    #     row_num += 1
    #     ws.write(row_num, 0, "Total", font_style_normal)
    #     for key, value in return_entry_sum.items():
    #         key = key.split('__')[0]
    #         ws.write(row_num, common.index(key), value or 0, font_style_normal)


    #     row_num += 2
    #     ws.write(row_num, 0, "Grand Total", font_style_bold)

    #     grand_total = data.get('grand_total')

    #     for key, value in grand_total.items():
    #         key = key.split('__')[0]
    #         ws.write(row_num, common.index(key), value or 0, font_style_bold)
    #     wb.save(response)
    #     return response

    def export_to_excel(self, data):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="purchase_book.xls"'

        # Add 'taxable_amount' to the common fields
        common = ['bill_date', "bill_no", "pp_no", "vendor_name", "vendor_pan", "amount", "taxable_amount", "tax_amount", "non_tax_purchase"]
        common.insert(0, 'idtblpurchaseEntry')
        extra = ["import","importCountry","importNumber", "importDate"]
        

        wb, ws, row_num, font_style_normal, font_style_bold = self.init_xls(
            "Purchase Book", common+extra
        )
        purchase_entry = data.get('purchase_entry')
        rows = purchase_entry.values_list(*common)

        for row in rows:
            # Add zeros for the extra fields (4 zeros because extra has 4 elements)
            row = row + (0, 0, 0, 0)
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style_normal)

        purchase_entry_sum = data.get('purchase_entry_sum')
        print(purchase_entry_sum)

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)
        for key, value in purchase_entry_sum.items():
            key = key.split('__')[0]
            # Make sure to handle the case where taxable_amount might not be in the sum
            if key in common:
                ws.write(row_num, common.index(key), value or 0, font_style_normal)

        common[0] = "idtblpurchaseReturn"
        columns2 = common+extra

        row_num += 1
        ws.write(row_num, 0, "")
        row_num += 1
        ws.write(row_num, 0, "Purchase Return", font_style_bold)
        row_num += 1

        new_columns = ["id"] + columns2[1:]
        for col_num in range(len(columns2)):
            ws.write(row_num, col_num, new_columns[col_num], font_style_bold)

        return_entry = data.get('return_entry')
        rows2 = return_entry.values_list(*common)
        return_entry_sum = data.get('return_entry_sum')

        for row in rows2:
            row = row + (0, 0, 0, 0)
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style_normal)

        row_num += 1
        ws.write(row_num, 0, "Total", font_style_normal)
        for key, value in return_entry_sum.items():
            key = key.split('__')[0]
            if key in common:
                ws.write(row_num, common.index(key), value or 0, font_style_normal)

        row_num += 2
        ws.write(row_num, 0, "Grand Total", font_style_bold)

        grand_total = data.get('grand_total')

        for key, value in grand_total.items():
            key = key.split('__')[0]
            if key in common:
                ws.write(row_num, common.index(key), value or 0, font_style_bold)
        
        wb.save(response)
        return response

    def get(self, request, *args, **kwargs):
        today = date.today()
        from_date = request.GET.get('fromDate', today)
        to_date = request.GET.get('toDate', today)
        format = request.GET.get('format', None)

        purchase_entry = TblpurchaseEntry.objects.filter(bill_date__range=[from_date, to_date])
        return_entry = TblpurchaseReturn.objects.filter(bill_date__range=[from_date, to_date])
        purchase_entry_sum = dict()
        return_entry_sum = dict()
        grand_total = dict()

        if purchase_entry:
            purchase_entry_sum = purchase_entry.aggregate(Sum('amount'), Sum('tax_amount'), Sum('non_tax_purchase'))
        if return_entry:
            return_entry_sum = return_entry.aggregate(Sum('amount'), Sum('tax_amount'), Sum('non_tax_purchase'))
            for key in purchase_entry_sum.keys():
                grand_total[key] = purchase_entry_sum[key] - return_entry_sum[key]

        context = {'purchase_entry':purchase_entry, 'return_entry':return_entry,
                    'purchase_entry_sum':purchase_entry_sum, 'return_entry_sum': return_entry_sum, 'grand_total': grand_total}
        
        if format and format =='xls':
            return self.export_to_excel(data=context)


        return render(request, 'purchase/purchase_book.html', context)


class VendorWisePurchaseView(IsAdminMixin, View):

    def get(self, request):
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        vendors = { v[0]:{'id':v[1], 'name':v[0], 'purchases':[]} for v in Vendor.objects.values_list('name', 'id')}
        if from_date and to_date:
            purchases = ''
        else:
            purchases = Purchase.objects.all()

        for purchase in purchases:
            vendor = vendors.get(purchase.vendor.name)
            vendor['purchases'].append(purchase)
            
        data = [i for i in vendors.values()]

        return render(request, 'purchase/vendorwisepurchase.html', {'object_list':data})



"""  ***************   Asset Purchase  ****************  """


from .models import AssetPurchase, Asset, AssetPurchaseItem
from .forms import AssetPurchaseForm

class AssetPurchaseMixin(IsAdminMixin):
    model = AssetPurchase
    form_class = AssetPurchaseForm
    paginate_by = 10
    queryset = AssetPurchase.objects.filter(status=True,is_deleted=False)
    success_url = reverse_lazy('assetpurchase_list')

class AssetPurchaseList(AssetPurchaseMixin, ListView):
    template_name = "assetpurchase/assetpurchase_list.html"
    queryset = AssetPurchase.objects.filter(status=True,is_deleted=False)

class AssetPurchaseDetail(AssetPurchaseMixin, DetailView):
    template_name = "assetpurchase/assetpurchase_detail.html"


class AssetPurchaseUpdate(AssetPurchaseMixin, UpdateView):
    template_name = "update.html"

# class AssetPurchaseDelete(AssetPurchaseMixin, DeleteMixin, View):
#     pass

# class AssetPurchaseCreate(IsAdminMixin, CreateView):
#     model = AssetPurchase
#     form_class = AssetPurchaseForm
#     template_name = "assetpurchase/assetpurchase_create.html"

#     def post(self, request):
#         bill_no = request.POST.get('bill_no', None)
#         bill_date = request.POST.get('bill_date', None)
#         vendor_id = request.POST.get('vendor')
#         sub_total = request.POST.get('sub_total')
#         discount_percentage = request.POST.get('discount_percentage')
#         discount_amount = request.POST.get('discount_amount')
#         taxable_amount = request.POST.get('taxable_amount')
#         non_taxable_amount = request.POST.get('non_taxable_amount')
#         tax_amount = request.POST.get('tax_amount')
#         grand_total = request.POST.get('grand_total')
#         amount_in_words = request.POST.get('amount_in_words')
#         payment_mode = request.POST.get('payment_mode')
#         debit_account = request.POST.get('debit_account', None)


#         vendor=None
#         try:
#             v_id = int(vendor_id)
#             vendor = Vendor.objects.get(pk=v_id)
#         except Exception as e:
#             vendor = Vendor.objects.create(name=vendor_id)
        
#         asset_purchase = AssetPurchase(
#             bill_no=bill_no,
#             vendor=vendor,sub_total=sub_total, bill_date=bill_date,
#             discount_percentage=discount_percentage,discount_amount=discount_amount,
#             taxable_amount=taxable_amount, non_taxable_amount=non_taxable_amount,
#             tax_amount=tax_amount, grand_total=grand_total,
#             amount_in_words=amount_in_words, payment_mode=payment_mode
#         )
#         asset_purchase.save()


#         selected_item_list = request.POST.get('select_items_list', [])
#         selected_item_list = selected_item_list.split(',')

#         debit_ledger = AccountLedger.objects.get(pk=int(debit_account))
#         depn_group, _ = AccountChart.objects.get_or_create(group='Depreciation')
#         depn_ledger, _ = AccountLedger.objects.get_or_create(account_chart=depn_group, ledger_name=f"{debit_ledger.ledger_name} Depreciation")
#         total_depreciation_amount = 0
#         for item in selected_item_list:
#             if not Asset.objects.filter(title=item).exists():
#                 depn = int(request.POST.get(f'id_depn_{item}'))
#                 asset = Asset.objects.create(title=item, depreciation_pool_id=int(depn))
#             else:
#                 asset = Asset.objects.get(title=item)
#             quantity = float(request.POST.get(f'id_bill_item_quantity_{item}'))
#             rate = float(request.POST.get(f'id_bill_item_rate_{item}'))
#             item_total = rate * quantity
#             item_purchased = AssetPurchaseItem.objects.create(asset=asset, asset_purchase=asset_purchase, rate=rate, quantity=quantity, item_total=item_total)

#             depreciation_amount, miti = calculate_depreciation(item_total, asset.depreciation_pool.percentage, bill_date)
#             depreciation_amount = decimal.Decimal(depreciation_amount)
#             net_amount = decimal.Decimal(item_total)-depreciation_amount

#             try:
#                 subled = AccountSubLedger.objects.get(sub_ledger_name=f'{asset.title}', ledger=debit_ledger)
#                 prev_value = subled.total_value
#                 subledgertracking = AccountSubLedgerTracking.objects.create(subledger = subled, prev_amount= subled.total_value)

#                 subled.total_value += net_amount
#                 subled.save()

#                 subledgertracking.new_amount=subled.total_value
#                 subledgertracking.value_changed = subled.total_value - prev_value
#                 subledgertracking.save()
#             except AccountSubLedger.DoesNotExist:
#                 subledger = AccountSubLedger.objects.create(sub_ledger_name=f'{asset.title} - Purchase', total_value= net_amount, ledger=debit_ledger)
#                 subledgertracking = AccountSubLedgerTracking.objects.create(subledger=subledger, new_amount=decimal.Decimal(net_amount), value_changed=decimal.Decimal(net_amount))

#             Depreciation.objects.create(item=item_purchased, miti=miti, depreciation_amount=depreciation_amount, net_amount=net_amount, ledger=debit_ledger)

#             try:
#                 sub_led = AccountSubLedger.objects.get(sub_ledger_name=f"{asset.title} Depreciation",ledger=depn_ledger)
#                 prev_value = subled.total_value
#                 subledgertracking = AccountSubLedgerTracking.objects.create(subledger = subled, prev_amount= subled.total_value)

#                 sub_led.total_value += depreciation_amount
#                 sub_led.save()

#                 subledgertracking.new_amount=subled.total_value
#                 subledgertracking.value_changed = subled.total_value - prev_value
#                 subledgertracking.save()
#             except AccountSubLedger.DoesNotExist:
#                 subledger = AccountSubLedger.objects.create(sub_ledger_name=f"{asset.title} Depreciation",ledger=depn_ledger,total_value=depreciation_amount)
#                 subledgertracking = AccountSubLedgerTracking.objects.create(subledger=subledger, new_amount=decimal.Decimal(net_amount), value_changed=decimal.Decimal(net_amount))

#             depn_ledger.total_value += depreciation_amount
#             total_depreciation_amount+= depreciation_amount
#             depn_ledger.save()

#         if payment_mode != 'Credit':
#             if debit_account:
#                 try:
#                     credit_ledger = AccountLedger.objects.get(ledger_name='Cash-In-Hand')
#                     journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username)

#                     grand_total = decimal.Decimal(grand_total)
#                     tax_amt = decimal.Decimal(tax_amount)

#                     total_debit_amt = grand_total - tax_amt
                    
#                     if tax_amt > 0:
#                         vat_receivable =  AccountLedger.objects.get(ledger_name='VAT Receivable')
#                         vat_receivable.total_value += tax_amt
#                         vat_receivable.save()
#                         TblDrJournalEntry.objects.create(ledger=vat_receivable, journal_entry=journal_entry, particulars=f'Vat receivable from {bill_no}', debit_amount=tax_amt)

#                     TblDrJournalEntry.objects.create(ledger=debit_ledger, journal_entry=journal_entry, particulars=f'Debit from bill {bill_no}', debit_amount=total_debit_amt)
#                     debit_ledger.total_value += total_debit_amt
#                     debit_ledger.save()
#                     TblCrJournalEntry.objects.create(ledger=credit_ledger, journal_entry=journal_entry,particulars=f'Cash cr. from bill {bill_no}', credit_amount=grand_total)
#                     credit_ledger.total_value -= grand_total
#                     credit_ledger.save()
#                     journal_entry.journal_total = total_debit_amt
#                     journal_entry.save()
#                 except Exception as e:
#                     print(e)
#         else:
#             if debit_account:
#                 vendor_name = vendor.name
#                 try:
#                     credit_ledger = None
#                     if not AccountLedger.objects.filter(ledger_name=vendor_name).exists():
#                         account_chart = AccountChart.objects.get(group='Sundry Creditors')
#                         credit_ledger = AccountLedger(ledger_name=vendor_name, account_chart=account_chart)
#                         credit_ledger.save()
                        
#                     credit_ledger = AccountLedger.objects.get(ledger_name=vendor_name)
#                     debit_ledger = AccountLedger.objects.get(pk=int(debit_account))
#                     journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username)

#                     grand_total = decimal.Decimal(grand_total)
#                     tax_amt = decimal.Decimal(tax_amount)

#                     total_debit_amt = grand_total - tax_amt
                    
#                     if tax_amt > 0:
#                         vat_receivable =  AccountLedger.objects.get(ledger_name='VAT Receivable')
#                         vat_receivable.total_value += tax_amt
#                         vat_receivable.save()
#                         TblDrJournalEntry.objects.create(ledger=vat_receivable, journal_entry=journal_entry, particulars=f'Vat receivable from {bill_no}', debit_amount=tax_amt)

#                     TblDrJournalEntry.objects.create(ledger=debit_ledger, journal_entry=journal_entry, particulars=f'Debit from bill {bill_no}', debit_amount=total_debit_amt)
#                     debit_ledger.total_value += total_debit_amt
#                     debit_ledger.save()
#                     TblCrJournalEntry.objects.create(ledger=credit_ledger, journal_entry=journal_entry,particulars=f'Cash cr. from bill {bill_no}', credit_amount=grand_total)
#                     credit_ledger.total_value += grand_total
#                     credit_ledger.save()

#                     journal_entry.journal_total = grand_total
#                     journal_entry.save()
#                 except Exception as e:
#                     print(e)

#         debit_ledger.total_value -= total_depreciation_amount
#         debit_ledger.save()

#         return redirect('/asset/')
   
 
 
from django.db import transaction
class AssetPurchaseCreate(IsAdminMixin, CreateView):
    model = AssetPurchase
    form_class = AssetPurchaseForm
    template_name = "assetpurchase/assetpurchase_create.html"
    @transaction.atomic()
    def post(self, request):
        print(request.POST)
        bill_no = request.POST.get('bill_no', None)
        bill_date = request.POST.get('bill_date', None)
        vendor_id = request.POST.get('vendor')
        sub_total = request.POST.get('sub_total')
        discount_percentage = request.POST.get('discount_percentage')
        discount_amount = request.POST.get('discount_amount')
        taxable_amount = request.POST.get('taxable_amount')
        non_taxable_amount = request.POST.get('non_taxable_amount')
        tax_amount = request.POST.get('tax_amount')
        grand_total = request.POST.get('grand_total')
        amount_in_words = request.POST.get('amount_in_words')
        payment_mode = request.POST.get('payment_mode')
        debit_account = request.POST.get('debit_account', None)


        vendor=None
        try:
            v_id = int(vendor_id)
            vendor = Vendor.objects.get(pk=v_id)
        except Exception as e:
            vendor = Vendor.objects.create(name=vendor_id)
        
        asset_purchase = AssetPurchase(
            bill_no=bill_no,
            vendor=vendor,sub_total=sub_total, bill_date=bill_date,
            discount_percentage=discount_percentage,discount_amount=discount_amount,
            taxable_amount=taxable_amount, non_taxable_amount=non_taxable_amount,
            tax_amount=tax_amount, grand_total=grand_total,
            amount_in_words=amount_in_words, payment_mode=payment_mode
        )
        asset_purchase.save()

        TblpurchaseEntry.objects.create(
            bill_no=bill_no, bill_date=bill_date, vendor_name=vendor.name, vendor_pan=vendor.pan_no,
            amount=grand_total, tax_amount=tax_amount, non_tax_purchase=non_taxable_amount, taxable_amount=taxable_amount
        )
        selected_item_list = request.POST.get('select_items_list', [])
        selected_item_list = selected_item_list.split(',')

        debit_ledger = AccountLedger.objects.get(pk=int(debit_account))
        depn_group, _ = AccountChart.objects.get_or_create(group='Depreciation')
        depn_ledger, _ = AccountLedger.objects.get_or_create(account_chart=depn_group, ledger_name=f"{debit_ledger.ledger_name} Depreciation")
        total_depreciation_amount = 0
        for item in selected_item_list:
            dashed_title = item.replace(' ', '-')
            if not Asset.objects.filter(title=item).exists():
                depn = int(request.POST.get(f'id_depn_{dashed_title}'))
                asset = Asset.objects.create(title=item, depreciation_pool_id=int(depn))
            else:
                asset = Asset.objects.get(title=item)
            quantity = float(request.POST.get(f'id_bill_item_quantity_{dashed_title}'))
            rate = float(request.POST.get(f'id_bill_item_rate_{dashed_title}'))
            item_total = rate * quantity
            item_purchased = AssetPurchaseItem.objects.create(asset=asset, asset_purchase=asset_purchase, rate=rate, quantity=quantity, item_total=item_total)

            depreciation_amount, miti = calculate_depreciation(item_total, asset.depreciation_pool.percentage, bill_date)
            depreciation_amount = decimal.Decimal(depreciation_amount)
            net_amount = decimal.Decimal(item_total)-depreciation_amount

            try:
                subled = AccountSubLedger.objects.get(sub_ledger_name=f'{asset.title}', ledger=debit_ledger)
                prev_value = subled.total_value
                subledgertracking = AccountSubLedgerTracking.objects.create(subledger = subled, prev_amount= subled.total_value)

                subled.total_value += net_amount
                subled.save()

                subledgertracking.new_amount=subled.total_value
                subledgertracking.value_changed = subled.total_value - prev_value
                subledgertracking.save()
            except AccountSubLedger.DoesNotExist:
                subledger = AccountSubLedger.objects.create(sub_ledger_name=f'{asset.title} - Purchase', total_value= net_amount, ledger=debit_ledger)
                subledgertracking = AccountSubLedgerTracking.objects.create(subledger=subledger, new_amount=decimal.Decimal(net_amount), value_changed=decimal.Decimal(net_amount))

            Depreciation.objects.create(item=item_purchased, miti=miti, depreciation_amount=depreciation_amount, net_amount=net_amount, ledger=debit_ledger)

            try:
                sub_led = AccountSubLedger.objects.get(sub_ledger_name=f"{asset.title} Depreciation",ledger=depn_ledger)
                prev_value = sub_led.total_value
                subledgertracking = AccountSubLedgerTracking.objects.create(subledger = sub_led, prev_amount= sub_led.total_value)

                sub_led.total_value += depreciation_amount
                sub_led.save()

                subledgertracking.new_amount=sub_led.total_value
                subledgertracking.value_changed = sub_led.total_value - prev_value
                subledgertracking.save()
            except AccountSubLedger.DoesNotExist:
                subledger = AccountSubLedger.objects.create(sub_ledger_name=f"{asset.title} Depreciation",ledger=depn_ledger,total_value=depreciation_amount)
                subledgertracking = AccountSubLedgerTracking.objects.create(subledger=subledger, new_amount=decimal.Decimal(net_amount), value_changed=decimal.Decimal(net_amount))

            depn_ledger.total_value += depreciation_amount
            total_depreciation_amount+= depreciation_amount
            depn_ledger.save()
            update_cumulative_ledger_bill(depn_ledger)
        if payment_mode != 'Credit':
            if debit_account:
                try:
                    credit_ledger = AccountLedger.objects.get(ledger_name='Cash-In-Hand')
                    journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username)

                    grand_total = decimal.Decimal(grand_total)
                    tax_amt = decimal.Decimal(tax_amount)

                    total_debit_amt = grand_total - tax_amt
                    
                    if tax_amt > 0:
                        vat_receivable =  AccountLedger.objects.get(ledger_name='VAT Receivable')
                        vat_receivable.total_value += tax_amt
                        vat_receivable.save()
                        update_cumulative_ledger_bill(vat_receivable)
                        TblDrJournalEntry.objects.create(ledger=vat_receivable, journal_entry=journal_entry, particulars=f'Vat receivable from {bill_no}', debit_amount=tax_amt)

                    TblDrJournalEntry.objects.create(ledger=debit_ledger, journal_entry=journal_entry, particulars=f'Debit from bill {bill_no}', debit_amount=total_debit_amt)
                    debit_ledger.total_value += total_debit_amt
                    debit_ledger.save()
                    update_cumulative_ledger_bill(debit_ledger)
                    TblCrJournalEntry.objects.create(ledger=credit_ledger, journal_entry=journal_entry,particulars=f'Cash cr. from bill {bill_no}', credit_amount=grand_total)
                    credit_ledger.total_value -= grand_total
                    credit_ledger.save()
                    update_cumulative_ledger_bill(credit_ledger)
                    journal_entry.journal_total = total_debit_amt
                    journal_entry.save()
                except Exception as e:
                    print(e)
        else:
            if debit_account:
                vendor_name = vendor.name
                try:
                    credit_ledger = None
                    if not AccountLedger.objects.filter(ledger_name=vendor_name).exists():
                        account_chart = AccountChart.objects.get(group='Sundry Creditors')
                        credit_ledger = AccountLedger(ledger_name=vendor_name, account_chart=account_chart)
                        credit_ledger.save()
                        update_cumulative_ledger_bill(credit_ledger)                        
                    credit_ledger = AccountLedger.objects.get(ledger_name=vendor_name)
                    debit_ledger = AccountLedger.objects.get(pk=int(debit_account))
                    journal_entry = TblJournalEntry.objects.create(employee_name=request.user.username)

                    grand_total = decimal.Decimal(grand_total)
                    tax_amt = decimal.Decimal(tax_amount)

                    total_debit_amt = grand_total - tax_amt
                    
                    if tax_amt > 0:
                        vat_receivable =  AccountLedger.objects.get(ledger_name='VAT Receivable')
                        vat_receivable.total_value += tax_amt
                        vat_receivable.save()
                        update_cumulative_ledger_bill(vat_receivable)
                        TblDrJournalEntry.objects.create(ledger=vat_receivable, journal_entry=journal_entry, particulars=f'Vat receivable from {bill_no}', debit_amount=tax_amt)

                    TblDrJournalEntry.objects.create(ledger=debit_ledger, journal_entry=journal_entry, particulars=f'Debit from bill {bill_no}', debit_amount=total_debit_amt)
                    debit_ledger.total_value += total_debit_amt
                    debit_ledger.save()
                    update_cumulative_ledger_bill(debit_ledger)
                    TblCrJournalEntry.objects.create(ledger=credit_ledger, journal_entry=journal_entry,particulars=f'Cash cr. from bill {bill_no}', credit_amount=grand_total)
                    credit_ledger.total_value += grand_total
                    credit_ledger.save()
                    update_cumulative_ledger_bill(credit_ledger)
                    journal_entry.journal_total = grand_total
                    journal_entry.save()
                except Exception as e:
                    print(e)

        debit_ledger.total_value -= total_depreciation_amount
        debit_ledger.save()
        update_cumulative_ledger_bill(debit_ledger)
        return redirect('/asset/')
   
   
from django.views import View
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.contrib import messages
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal
import nepali_datetime

from purchase.models import TblpurchaseEntry


class PurchaseEntryUploadView(View):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            messages.error(request, "No file uploaded.", extra_tags='danger')
            return redirect(reverse_lazy('tblpurchaseentry_list'))

        wb = load_workbook(file)
        errors = []

        def clean_decimal(value):
            try:
                return Decimal(str(value).replace(",", "").strip()) if value else Decimal(0)
            except Exception:
                raise ValueError(f"Invalid decimal value: {value}")

        def parse_excel_date(miti_raw, row_idx):
            if isinstance(miti_raw, datetime):
                d = miti_raw.date()
                if d.year >= 2070:
                    bs_date = nepali_datetime.date(d.year, d.month, d.day)
                    return bs_date.to_datetime_date(), f"{d.year}/{d.month}/{d.day}"
                else:
                    raise ValueError(f"Row {row_idx}: Expected B.S. date, got A.D.")
            elif isinstance(miti_raw, str):
                if miti_raw.lower() in ['b.s.', 'miti', 'date']:
                    raise ValueError(f"Row {row_idx}: Skipped header row")
                try:
                    if '-' in miti_raw:
                        y, m, d = map(int, miti_raw.split('-'))
                    else:
                        y, m, d = map(int, miti_raw.split('/'))
                    bs_date = nepali_datetime.date(y, m, d)
                    return bs_date.to_datetime_date(), miti_raw
                except Exception:
                    raise ValueError(f"Row {row_idx}: Invalid B.S. date format: {miti_raw}")
            else:
                raise ValueError(f"Row {row_idx}: Unsupported date type: {type(miti_raw)}")

        for sheet in wb.worksheets:
            for idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
                try:
                    # B.S. Date and A.D. conversion
                    miti_raw = row[0].value
                    try:
                        ad_date, bs_miti = parse_excel_date(miti_raw, idx)
                    except Exception as e:
                        errors.append(str(e))
                        continue

                    bill_no = str(row[1].value).strip() if row[1].value else ''
                    vendor_name = str(row[2].value).strip() if row[2].value else ''
                    vendor_pan = str(row[3].value).strip() if row[3].value else ''

                    # Taxable Purchase
                    taxable_purchase = clean_decimal(row[6].value)
                    taxable_purchase_tax = clean_decimal(row[7].value)

                    # Taxable Import
                    taxable_import = clean_decimal(row[8].value)
                    taxable_import_tax = clean_decimal(row[9].value)

                    # Capital Purchase or Import and Tax (combined in last column)
                    capital_purchase = clean_decimal(row[10].value)  
                    capital_purchase_tax = clean_decimal(row[11].value) # last column (assumed tax included)

                    # Calculate total amount = all taxable + tax values
                    total_amount = (
                        taxable_purchase + taxable_purchase_tax +
                        taxable_import + taxable_import_tax +
                        capital_purchase_tax + capital_purchase
                    )

                    tax_amount = (
                        taxable_purchase_tax +
                        taxable_import_tax + capital_purchase_tax
                        # Note: If tax is separated in capital_purchase_tax, split accordingly
                    )

                    TblpurchaseEntry.objects.create(
                        bill_date=str(ad_date),
                        bill_no=bill_no,
                        vendor_name=vendor_name,
                        vendor_pan=vendor_pan,
                        amount=total_amount,
                        tax_amount=tax_amount,
                        # Optional: add if needed
                        non_tax_purchase=clean_decimal(row[5].value)  # Tax Free column
                    )

                    print(f"Row {idx}: Purchase entry saved successfully.")

                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    print(f"Row {idx} error: {str(e)}")

        if errors:
            messages.error(request, f"Some rows failed:\n{errors}", extra_tags='danger')
        else:
            messages.success(request, "Purchase entries uploaded successfully.", extra_tags='success')

        return redirect(reverse_lazy('purchase_book_list'))



from django.views import View
from django.shortcuts import render
from datetime import datetime
from django.db.models import Sum, F
from .models import ProductPurchase, Purchase

# class PurchasedProducts(View):
#     def get(self, request):
#         today_date = datetime.now().date()

#         purchases = Purchase.objects.filter(bill_date=today_date)
#         items = ProductPurchase.objects.filter(purchase__in=purchases)

#         # Grouping by category and product
#         grouped = items.values(
#             category_id=F("product__category__id"),
#             category_title=F("product__category__title"),
#             product_title=F("product__title"),
#             product_unit=F("product__unit"),
#             product_rate=F("rate")
#         ).annotate(
#             total_quantity=Sum("quantity"),
#             total_amount=Sum("item_total")
#         ).order_by("category_title")

#         # Organize data category-wise
#         category_data = {}
#         grand_total_quantity = 0
#         grand_total_amount = 0

#         for item in grouped:
#             cat_title = item["category_title"]
#             if cat_title not in category_data:
#                 category_data[cat_title] = {
#                     "products": [],
#                     "category_total_quantity": 0,
#                     "category_total_amount": 0
#                 }

#             category_data[cat_title]["products"].append(item)
#             category_data[cat_title]["category_total_quantity"] += item["total_quantity"]
#             category_data[cat_title]["category_total_amount"] += item["total_amount"]

#             grand_total_quantity += item["total_quantity"]
#             grand_total_amount += item["total_amount"]

#         data = {
#             "date": today_date,
#             "category_data": category_data,
#             "grand_total_quantity": grand_total_quantity,
#             "grand_total_amount": grand_total_amount,
#             "org" : Organization.objects.last()
#         }

#         return render(request, 'purchase/purchase_items_report.html', data)

from django.utils.dateparse import parse_date
class PurchasedProducts(View):
    def get(self, request):
        # Get date range from query params or use today
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')

        if from_date and to_date:
            from_date = parse_date(from_date)
            to_date = parse_date(to_date)
        else:
            from_date = to_date = datetime.now().date()

        # Filter purchases in date range
        purchases = Purchase.objects.filter(bill_date__range=(from_date, to_date))
        items = ProductPurchase.objects.filter(purchase__in=purchases)

        # Group by category and product
        grouped = items.values(
            category_id=F("product__category__id"),
            category_title=F("product__category__title"),
            product_title=F("product__title"),
            product_unit=F("product__unit"),
            product_rate=F("rate")
        ).annotate(
            total_quantity=Sum("quantity"),
            total_amount=Sum("item_total")
        ).order_by("category_title")

        # Organize data category-wise
        category_data = {}
        grand_total_quantity = 0
        grand_total_amount = 0

        for item in grouped:
            cat_title = item["category_title"]
            if cat_title not in category_data:
                category_data[cat_title] = {
                    "products": [],
                    "category_total_quantity": 0,
                    "category_total_amount": 0
                }

            category_data[cat_title]["products"].append(item)
            category_data[cat_title]["category_total_quantity"] += item["total_quantity"]
            category_data[cat_title]["category_total_amount"] += item["total_amount"]

            grand_total_quantity += item["total_quantity"]
            grand_total_amount += item["total_amount"]

        data = {
            "date": f"{from_date} to {to_date}" if from_date != to_date else str(from_date),
            "category_data": category_data,
            "grand_total_quantity": grand_total_quantity,
            "grand_total_amount": grand_total_amount,
            "org": Organization.objects.last(),
            "request": request  # needed to show input values in the template
        }

        return render(request, 'purchase/purchase_items_report.html', data)
        
from .models import ImportProductPurchase, ImportPurchase
from .forms import ImportProductPurchaseForm
class ImportProductPurchaseCreateView(IsAdminMixin, CreateView):
    model = ImportProductPurchase
    form_class = ImportProductPurchaseForm
    template_name = "importpurchase/importpurchase_create.html"

    def create_subledgers(self, product, item_total, debit_account):
        debit_account = get_object_or_404(AccountLedger, pk=int(debit_account))
        subledgername = f'{product.title} ({product.category.title}) - Purchase'
        try:
            sub = AccountSubLedger.objects.get(sub_ledger_name=subledgername, ledger=debit_account)
            prev_value = sub.total_value
            subledgertracking = AccountSubLedgerTracking.objects.create(subledger = sub, prev_amount= prev_value)
            sub.total_value += decimal.Decimal(item_total)
            sub.save()
            subledgertracking.new_amount=sub.total_value
            subledgertracking.value_changed = sub.total_value - prev_value
            subledgertracking.save()
        except AccountSubLedger.DoesNotExist:
            subledger = AccountSubLedger.objects.create(sub_ledger_name=subledgername, ledger=debit_account, total_value=item_total)
            subledgertracking = AccountSubLedgerTracking.objects.create(subledger=subledger, new_amount=decimal.Decimal(item_total), value_changed=decimal.Decimal(item_total))

    def create_accounting_multiple_ledger(self, debit_account_id, payment_mode:str, username:str, sub_total, tax_amount, vendor, excise_duty_amount):
        sub_total = decimal.Decimal(sub_total)
        tax_amount = decimal.Decimal(tax_amount)
        total_amount =  sub_total+ tax_amount + excise_duty_amount

        cash_ledger = get_object_or_404(AccountLedger, ledger_name='Cash-In-Hand')
        vat_receivable = get_object_or_404(AccountLedger, ledger_name='VAT Receivable')
        excise_duty_receivable = get_object_or_404(AccountLedger, ledger_name='Excise Duty Receivable')
        debit_account = get_object_or_404(AccountLedger, pk=int(debit_account_id))
        
        journal_entry = TblJournalEntry.objects.create(employee_name=username, journal_total = total_amount)
        TblDrJournalEntry.objects.create(journal_entry=journal_entry, particulars=f"Automatic: {debit_account.ledger_name} A/c Dr.", debit_amount=sub_total, ledger=debit_account)
        debit_account.total_value += sub_total
        debit_account.save()
        update_cumulative_ledger_bill(debit_account)
        if tax_amount > 0:
            TblDrJournalEntry.objects.create(journal_entry=journal_entry, particulars="Automatic: VAT Receivable A/c Dr.", debit_amount=tax_amount, ledger=vat_receivable)
            vat_receivable.total_value += tax_amount
            vat_receivable.save()
            update_cumulative_ledger_bill(vat_receivable)
        if excise_duty_amount > 0:
            TblDrJournalEntry.objects.create(journal_entry=journal_entry, particulars="Automatic: Excise Duty Receivable A/c Dr.", debit_amount=excise_duty_amount, ledger=excise_duty_receivable)
            excise_duty_receivable.total_value += excise_duty_amount
            excise_duty_receivable.save()
            update_cumulative_ledger_bill(excise_duty_receivable)
        if payment_mode.lower().strip() == "credit":
            try:
                vendor_ledger = AccountLedger.objects.get(ledger_name=vendor)
                vendor_ledger.total_value += total_amount
                vendor_ledger.save()
                update_cumulative_ledger_bill(vendor_ledger)
            except AccountLedger.DoesNotExist:
                chart = AccountChart.objects.get(group__iexact='Sundry Creditors')
                vendor_ledger = AccountLedger.objects.create(ledger_name=vendor, total_value=total_amount, is_editable=True, account_chart=chart)
                create_cumulative_ledger_bill(vendor_ledger)
            TblCrJournalEntry.objects.create(journal_entry=journal_entry, particulars=f"Automatic: To {vendor_ledger.ledger_name} A/c", credit_amount=total_amount, ledger=vendor_ledger)
        else:
            TblCrJournalEntry.objects.create(journal_entry=journal_entry, particulars=f"Automatic: To {cash_ledger.ledger_name} A/c", credit_amount=total_amount, ledger=cash_ledger)
            cash_ledger.total_value -= total_amount
            cash_ledger.save()
            update_cumulative_ledger_bill(cash_ledger)

    def create_accounting_single_ledger(self, ledger_totals, payment_mode: str, username: str, 
                                    tax_amount: decimal.Decimal, excise_duty_amount: decimal.Decimal, 
                                    vendor: str, importtax_amount: decimal.Decimal):
        """
        Create a single journal entry for all product ledgers
        ledger_totals: dict of {ledger_id: total_amount}
        """

        
        # Calculate grand total
        sub_total = sum(decimal.Decimal(total) for total in ledger_totals.values())
        total_amount = sub_total + tax_amount + excise_duty_amount + importtax_amount

        # Get required ledgers
        cash_ledger = get_object_or_404(AccountLedger, ledger_name='Cash-In-Hand')
        vat_receivable = get_object_or_404(AccountLedger, ledger_name='VAT Receivable')
        import_tax_receivable = get_object_or_404(AccountLedger, ledger_name='Import Tax Receivable')
        excise_duty_receivable = get_object_or_404(AccountLedger, ledger_name='Excise Duty Receivable')
        
        # Create single journal entry
        journal_entry = TblJournalEntry.objects.create(
            employee_name=username, 
            journal_total=total_amount
        )
        
        # Create debit entries for each product ledger
        for ledger_id, amount in ledger_totals.items():
            ledger = get_object_or_404(AccountLedger, pk=int(ledger_id))
            TblDrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: {ledger.ledger_name} A/c Dr.",
                debit_amount=decimal.Decimal(amount),
                ledger=ledger
            )
            # Update ledger balance
            ledger.total_value += decimal.Decimal(amount)
            ledger.save()
            update_cumulative_ledger_bill(ledger)
        
        # Add tax and excise duty entries if applicable
        if tax_amount > 0:
            TblDrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars="Automatic: VAT Receivable A/c Dr.",
                debit_amount=tax_amount,
                ledger=vat_receivable
            )
            vat_receivable.total_value += tax_amount
            vat_receivable.save()
            update_cumulative_ledger_bill(vat_receivable)
        
        if excise_duty_amount > 0:
            TblDrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars="Automatic: Excise Duty Receivable A/c Dr.",
                debit_amount=excise_duty_amount,
                ledger=excise_duty_receivable
            )
            excise_duty_receivable.total_value += excise_duty_amount
            excise_duty_receivable.save()
            update_cumulative_ledger_bill(excise_duty_receivable)

        if importtax_amount > 0:
            TblDrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars="Automatic: Import Tax Receivable A/c Dr.",
                debit_amount=importtax_amount,
                ledger=import_tax_receivable
            )
            import_tax_receivable.total_value += importtax_amount
            import_tax_receivable.save()
            update_cumulative_ledger_bill(import_tax_receivable)
        
        # Create credit entry
        if payment_mode.lower().strip() == "credit":
            try:
                vendor_ledger = AccountLedger.objects.get(ledger_name=vendor)
                vendor_ledger.total_value += total_amount
                vendor_ledger.save()
                update_cumulative_ledger_bill(vendor_ledger)
            except AccountLedger.DoesNotExist:
                chart = AccountChart.objects.get(group__iexact='Sundry Creditors')
                vendor_ledger = AccountLedger.objects.create(
                    ledger_name=vendor,
                    total_value=total_amount,
                    is_editable=True,
                    account_chart=chart
                )
                create_cumulative_ledger_bill(vendor_ledger)
            

            
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {vendor_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=vendor_ledger
            )
        elif payment_mode.lower().strip() == "mobile payment":
            # Get required ledgers
            mobile_payment_ledger = get_object_or_404(AccountLedger, ledger_name='Mobile Payments')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {mobile_payment_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=mobile_payment_ledger
            )
            mobile_payment_ledger.total_value -= total_amount
            mobile_payment_ledger.save()
            update_cumulative_ledger_bill(mobile_payment_ledger)
        elif payment_mode.lower().strip() == "credit card":
            # Get required ledgers
            credit_card_ledger = get_object_or_404(AccountLedger, ledger_name='Card Transactions')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {credit_card_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=credit_card_ledger
            )
            credit_card_ledger.total_value -= total_amount
            credit_card_ledger.save()
            update_cumulative_ledger_bill(credit_card_ledger)
        elif payment_mode.lower().strip() == "complimentary":
            # Get required ledgers
            complimentary_expense_ledger = get_object_or_404(AccountLedger, ledger_name='Complimentary Expenses')
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {complimentary_expense_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=complimentary_expense_ledger
            )
            complimentary_expense_ledger.total_value -= total_amount
            complimentary_expense_ledger.save()
            update_cumulative_ledger_bill(complimentary_expense_ledger)
        else:
            TblCrJournalEntry.objects.create(
                journal_entry=journal_entry,
                particulars=f"Automatic: To {cash_ledger.ledger_name} A/c",
                credit_amount=total_amount,
                ledger=cash_ledger
            )
            cash_ledger.total_value -= total_amount
            cash_ledger.save()
            update_cumulative_ledger_bill(cash_ledger)

    def form_invalid(self, form) -> HttpResponse:
        return self.form_valid(form)
    @transaction.atomic()    
    def form_valid(self, form):
        form_data = form.data 
        print(form.data)
        bill_no = form_data.get('bill_no', None)
        bill_date = form_data.get('bill_date', None)
        pp_no = form_data.get('pp_no',None)
        vendor_id = form_data.get('vendor')
        sub_total = form_data.get('sub_total')
        discount_percentage = form_data.get('discount_percentage')
        discount_amount = form_data.get('discount_amount')
        taxable_amount = form_data.get('taxable_amount')
        non_taxable_amount = form_data.get('non_taxable_amount')
        tax_amount = form_data.get('tax_amount')
        grand_total = form_data.get('grand_total')
        amount_in_words = form_data.get('amount_in_words')
        payment_mode = form_data.get('payment_mode')
        debit_account = form_data.get('debit_account')
        excise_duty_amount = form_data.get('excise_duty_amount', 0.0)
        importtax_amount = form_data.get('importtax_amount')
        # print(debit_account)
        purchase_object = ImportPurchase(
            bill_no=bill_no,
            vendor_id=vendor_id,sub_total=sub_total, bill_date=bill_date,
            discount_percentage=discount_percentage,discount_amount=discount_amount,
            taxable_amount=taxable_amount, non_taxable_amount=non_taxable_amount,
            tax_amount=tax_amount, grand_total=grand_total,
            amount_in_words=amount_in_words, payment_mode=payment_mode,
            excise_duty_amount=decimal.Decimal(excise_duty_amount), import_tax_amount = decimal.Decimal(importtax_amount)
        )
        purchase_object.save()

        product_ids =  form_data.get('product_id_list', '')
        product_taxable_info = form_data.get('product_taxable_info', '')
        product_ledger_info = form_data.get('ledger_id_list', '')
        product_ledger_info_parse = json.loads(product_ledger_info)
        # print(product_ledger_info)
        no_of_items_sent = len(product_ledger_info_parse)
        product_category_info = form_data.get('product_category_info')
        print(product_category_info)


        new_items_name = {}
        new_product_categories = {}
        new_product_ledgers = {}
        if product_taxable_info and len(product_taxable_info) > 0:
            new_items_name = json.loads(product_taxable_info)
            # print(new_items_name)


            new_product_categories = json.loads(product_category_info)
            new_product_ledgers = json.loads(product_ledger_info)

        item_name = ''

        total_quantity = 0
        vendor = Vendor.objects.get(pk=vendor_id)
        vendor_name = vendor.name
        vendor_pan = vendor.pan_no

        if product_ids:
            product_ids = product_ids.split(',')


        
        if product_ledger_info and len(product_ledger_info) > 0:
            product_ledgers = json.loads(product_ledger_info)
            
            for product_id, ledger_info in product_ledgers.items():
                try:
                    product_id = int(product_id)
                    ledger_id = int(ledger_info['ledgerId'])
                    total = float(ledger_info['total'])
                    # print(product_id)
                    # print(ledger_id)
                    
                    quantity = float(form_data.get(f'id_bill_item_quantity_{product_id}'))
                    rate = float(form_data.get(f'id_bill_item_rate_{product_id}'))
                    importtax_product_percent = float(form_data.get(f'id_bill_item_importtaxpercent_{product_id}'))
                    item_total = quantity * rate
                    # print(quantity)
                    # print(rate)

                    # Get the product and ledger objects
                    prod = Product.objects.get(pk=product_id)
                    ledger = AccountLedger.objects.get(pk=ledger_id)

                    # Debit the ledger for the product
                    self.create_subledgers(prod, item_total, ledger_id)
                    ImportProductPurchase.objects.create(product=prod, purchase=purchase_object, quantity=quantity, rate=rate, item_total=total, importtax_percent=importtax_product_percent)
                    from organization.models import Branch
                    from product.models import BranchStock
                    BranchStock.objects.create(branch=Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first(), product=prod, quantity=quantity)
                except (ValueError, Product.DoesNotExist, AccountLedger.DoesNotExist):
                    pass
        


        if new_items_name:
            for k, v in new_items_name.items():
                category_name = new_product_categories.get(k, '').lower().strip()
                # print(category_name)
                # ledger_name = new_product_ledgers
                if ProductCategory.objects.filter(title__iexact=category_name).exists():
                    category = ProductCategory.objects.filter(title__iexact=category_name).first()
                else:
                    try:
                        ProductCategory.objects.create(title=category_name)
                    except IntegrityError:
                        pass
                category = ProductCategory.objects.filter(title__iexact=category_name).first()
                rate = float(form_data.get(f'id_bill_item_rate_{k}'))
                quantity = float(form_data.get(f'id_bill_item_quantity_{k}'))
                importtax_product_percent = float(form_data.get(f'id_bill_item_importtaxpercent_{k}'))
                item_total = quantity * rate
                is_taxable = True if (v == "true" or v == True) else False
                ledger_info = json.loads(product_ledger_info)
                # print(ledger_info)
                ledger_id = int(ledger_info.get(k, {}).get('ledgerId', ''))
                ledger = AccountLedger.objects.get(id=ledger_id)
                # print(ledger)
                clean_title = k.replace('-', ' ')
                try:
                    prod = Product.objects.create(category=category, title=clean_title, is_taxable=is_taxable, price=rate, ledger=ledger, is_billing_item = False, importtax_percent=importtax_product_percent, excise_duty_applicable=False)
                except IntegrityError:
                    prod = Product.objects.get(title__iexact=clean_title)
                self.create_subledgers(prod, item_total, ledger_id)
                ImportProductPurchase.objects.create(product=prod, purchase=purchase_object, quantity=quantity, rate=rate, item_total=item_total, importtax_percent=importtax_product_percent)
                from organization.models import Branch
                from product.models import BranchStock
                BranchStock.objects.create(branch=Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first(), product=prod, quantity=quantity)
        TblpurchaseEntry.objects.create(
            bill_no=bill_no, bill_date=bill_date, pp_no=pp_no, vendor_name=vendor_name, vendor_pan=vendor_pan,
            item_name=item_name, quantity=total_quantity, amount=grand_total, tax_amount=tax_amount, taxable_amount=taxable_amount, non_tax_purchase=non_taxable_amount, excise_duty_amount=decimal.Decimal(excise_duty_amount), import_tax_amount=decimal.Decimal(importtax_amount), purchase_id=purchase_object.id
        )
        vendor_detail = str(vendor.pk)+' '+ vendor_name
        # self.create_accounting(debit_account_id=debit_account, payment_mode=payment_mode, username=self.request.user.username, sub_total=sub_total, tax_amount=tax_amount, vendor=vendor_detail)
        sub_tax = decimal.Decimal(tax_amount)
        fraction_tax = sub_tax/no_of_items_sent
        sub_excise_duty = decimal.Decimal(excise_duty_amount)
        fraction_excise_duty = sub_excise_duty/no_of_items_sent
        print(fraction_tax)
        ledger_totals = {}
        if product_ledger_info and len(product_ledger_info) > 0:
            product_ledgers = json.loads(product_ledger_info)
            for product_id, ledger_info in product_ledgers.items():
                ledger_id = ledger_info['ledgerId']
                total = float(ledger_info['total'])
                if ledger_id in ledger_totals:
                    ledger_totals[ledger_id] += total
                else:
                    ledger_totals[ledger_id] = total
        
        # Create single journal entry with all ledgers
        if ledger_totals:
            print("This is ledger totals", ledger_totals)
            self.create_accounting_single_ledger(
                ledger_totals=ledger_totals,
                payment_mode=payment_mode,
                username=self.request.user.username,
                tax_amount=decimal.Decimal(tax_amount),
                excise_duty_amount=decimal.Decimal(excise_duty_amount),
                vendor=vendor_detail,
                importtax_amount = decimal.Decimal(importtax_amount)
            )

        return redirect('/importpur_chase/')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Fetch ledgers with the account_chart of purchases and expenses
        purchases_and_expenses_ledgers = AccountLedger.objects.filter(
            Q(account_chart__account_type='Expense') | Q(account_chart__group='Purchases')
        )
        # print(purchases_and_expenses_ledgers)

        # Add the fetched ledgers to the context
        context['purchases_and_expenses_ledgers'] = purchases_and_expenses_ledgers

        return context


class ImportPurchaseListView(IsAdminMixin, ListView):
    model = ImportPurchase
    queryset = ImportPurchase.objects.filter(is_deleted=False)
    template_name = 'importpurchase/importpurchase_list.html'


class ImportPurchaseDetailView(IsAdminMixin, DetailView):
    template_name = 'importpurchase/importpurchase_detail.html'
    queryset = ImportPurchase.objects.filter(is_deleted=False)

    def get_context_data(self, **kwargs):
        org = Organization.objects.first()
        context =  super().get_context_data(**kwargs)
        context['organization'] = org
        return context


from django.db import transaction
class ImportMarkPurchaseVoid(IsAdminMixin, View):
    @transaction.atomic()
    def post(self, request, *args, **kwargs):
        id = self.kwargs.get('pk')
        reason = request.POST.get('voidReason')
        purchase = get_object_or_404(ImportPurchase, pk=id)
        purchase.status = False
        purchase.save()

        purchased_products = purchase.productpurchase_set.all()

        org = Organization.objects.first()
        allow_negative_sales = org.allow_negative_sales
        
        from organization.models import Branch
        from product.models import BranchStock
        
        central_branch = Branch.objects.filter(is_central_billing=True, status=True, is_deleted=False).first()
        print(purchased_products)
        for product_purchase in purchased_products:
            quantity = product_purchase.quantity
            product = product_purchase.product
            
            # Get all stock entries for this product in central branch, ordered by quantity (ascending)
            branchstock_entries = BranchStock.objects.filter(
                    branch=central_branch, 
                    product=product, 
                    is_deleted=False
                ).order_by('quantity')
                
            remaining_quantity = quantity
                
                 
            for branchstock in branchstock_entries:
                if remaining_quantity <= 0:
                    break
                        
                available_quantity = branchstock.quantity
                    
                if remaining_quantity >= available_quantity:
                    # Delete the entry if quantity becomes zero
                    branchstock.quantity = 0
                    branchstock.is_deleted = True
                    branchstock.save()
                    remaining_quantity -= available_quantity
                else:
                    # Reduce the quantity
                    branchstock.quantity -= remaining_quantity
                    branchstock.save()
                    remaining_quantity = 0

        entry_obj = TblpurchaseEntry.objects.get(purchase_id=id)
        TblpurchaseReturn.objects.create(
            bill_date=entry_obj.bill_date,
            bill_no=entry_obj.bill_no,
            pp_no=entry_obj.pp_no,
            vendor_name=entry_obj.vendor_name,
            vendor_pan=entry_obj.vendor_pan,
            item_name=entry_obj.item_name,
            quantity=entry_obj.quantity,
            unit=entry_obj.unit,
            amount=entry_obj.amount,
            tax_amount=entry_obj.tax_amount,
            taxable_amount=entry_obj.taxable_amount,
            non_tax_purchase=entry_obj.non_tax_purchase,
            reason=reason,
            excise_duty_amount=entry_obj.excise_duty_amount,
            import_tax_amount=entry_obj.import_tax_amount,
        )
        
        return redirect(reverse_lazy("importpur_chase_detail", kwargs={"pk": id}))
